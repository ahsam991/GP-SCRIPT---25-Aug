"""POS Daily Report - GUI (Download Only / Download & Send).

Reuses the exact SQL query and file-generation logic of the old script
(pos_daily_report_automation.py) so visit data stays accurate, and adds the
circle-wise recipient mapping + separate per-circle emails from the new script.
"""

import os
import re
import smtplib
import threading
import warnings
import zipfile
from datetime import datetime

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import create_engine, text

warnings.filterwarnings("ignore", category=UserWarning, module="xlsxwriter.worksheet")

DB_PARAMS = {
    "dbname": "gp_dev",
    "user": "report_user",
    "password": "report#Gp*User!#__D",
    "host": "gp-stg.cf44ysgum7u8.ap-southeast-1.rds.amazonaws.com",
    "port": 5432  # Adjust if necessary
}


DATABASE_URL = (
    f"postgresql://{DB_PARAMS['user']}:{DB_PARAMS['password']}"
    f"@{DB_PARAMS['host']}:{DB_PARAMS['port']}/{DB_PARAMS['dbname']}"
)
engine = create_engine(DATABASE_URL)

def pos_daily_rpt(custom_date):
    return text(f"""
                WITH mv_data AS (
    SELECT
        *
    FROM (
        SELECT
            mvc.id visit_id,
--          mvc.user_id,
            mvc.contacted_by AS me,
            mvc.visit_date,
            r.route_no AS route,
            o.outlet_code,
            o.geo_checked_by,
            o.is_skitto,
            '(' || mvc.lat || ', ' || mvc.long || ')' visit_lat_long,
            mvc.operational_status,
            mvc.scheduled_visit,
            RANK() OVER (PARTITION BY o.outlet_code,mvc.user_id ORDER BY mvc.id DESC)
            rank
        FROM
            gp.market_visit_contacts mvc,
            gp.outlets o,
            gp.routes r
        WHERE
            mvc.location_id = o.id
            AND mvc.location_source_id = r.id
            AND mvc.visit_date = '{custom_date}') x
    WHERE
        rank = 1
        AND x.route not like '%V2TECH%'
),
user_data AS (WITH loc_data AS (
SELECT
    c. "name" AS "Circle",
    r. "name" AS "Region",
    cl. "name" AS "Cluster",
    tr. "name" AS "Territory",
    dh. "name" AS "Distribution House",
    dh.dh_code AS "DH Code",
    '(' || dh.latitude || ', ' || dh.longitude || ')' AS "DH Lat-Long",
    h.id AS dh_id
FROM
    gp.locations c,
    gp.locations r,
    gp.locations cl,
    gp.locations tr,
    gp.locations h,
    gp.house dh
WHERE
    c. "type" = 1
    AND r. "type" = 2
    AND cl. "type" = 3
    AND tr. "type" = 4
    AND h. "type" = 5
    AND r.parent = c.id
    AND cl.parent = r.id
    AND tr.parent = cl.id
    AND h.parent = tr.id
    AND h.is_deleted IS FALSE
    AND h.id = dh.location_id
    AND h. "name" NOT LIKE '%TEST'
)
SELECT
    ld.*,
    u.id AS user_id,
    ui.full_name AS "ME Name",
    ui.official_contact AS "ME Contact No.",
    u.email AS "ME Email",
    b.bundle_code AS "ME Code"
FROM
    gp.users u,
    gp.user_infos ui,
    gp.users_bundle_maps ubm,
    gp.bundles b,
    gp.bundles_house_maps bhm,
    loc_data ld
WHERE
    u.id = ubm.user_id
    AND ui.user_id = u.id
    AND ubm.bundle_id = b.id
    and bhm.bundle_id = b.id
    AND ld.dh_id = bhm.house_id
    AND '{custom_date}' BETWEEN ubm.from_date AND COALESCE(ubm.to_date, '{custom_date}')
      AND '{custom_date}' BETWEEN bhm.from_date AND COALESCE(bhm.to_date, '{custom_date}')
    AND b.bundle_code NOT LIKE '%TEST%'
),
survey_data AS (
SELECT 
    visit_id,
    MAX(CASE WHEN question = 'temporarily_&_permanently_closed_photo' THEN 
        CASE WHEN answer IS NULL OR answer = '' THEN '' 
        ELSE CONCAT('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mvcf.visit_date, 'YYYYMM'),'/', answer) END 
    END) AS "temporarily_&_permanently_closed_photo",
    MAX(CASE WHEN question = 'pre_execution_photo' THEN answer END) AS pre_execution_photo,
    MAX(CASE WHEN question = 'post_execution_photo' THEN answer END) AS post_execution_photo,
    MAX(CASE WHEN question = 'posm_exists' THEN answer END) AS posm_exists,
    MAX(CASE WHEN question = 'not_found_&_moved_photo' THEN 
        CASE WHEN answer IS NULL OR answer = '' THEN '' 
        ELSE CONCAT('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mvcf.visit_date, 'YYYYMM'),'/', answer) END 
    END) AS "not_found_&_moved_photo",
    MAX(CASE WHEN question = 'old_posm_list' THEN answer END) AS old_posm_list,
    MAX(CASE WHEN question = 'do_posm_remove' THEN answer END) AS do_posm_remove,
    MAX(CASE WHEN question = 'old_posm_remove_list' THEN answer END) AS old_posm_remove_list,
    MAX(CASE WHEN question = 'removed_posm_counts' THEN answer END) AS removed_posm_counts,
    MAX(CASE WHEN question = 'new_posm_selection' THEN answer END) AS new_posm_selection,
    MAX(CASE WHEN question = 'new_posm_counts' THEN answer END) AS new_posm_counts, 
--    MAX(CASE WHEN question = 'fascia_type' THEN answer END) AS fascia_type,
--    MAX(CASE WHEN question = 'fascia' THEN answer END) AS fascia,
    MAX(CASE WHEN question = 'execution_photo_right' THEN answer END) AS execution_photo_right,
    MAX(CASE WHEN question = 'execution_photo_left' THEN answer END) AS execution_photo_left,
    MAX(CASE WHEN question = 'execution_photo_center' THEN answer END) AS execution_photo_center,
        MAX(CASE WHEN question = 'which_is_right_for_pos' THEN answer END) AS which_is_right_for_pos,
    MAX(CASE WHEN question = 'pos_status' THEN answer END) AS pos_status,
    MAX(CASE WHEN question = 'pos_structure' THEN answer END) AS pos_structure,
    MAX(CASE WHEN question = 'pos_location' THEN answer END) AS pos_location,
    MAX(CASE WHEN question = 'pos_business_type' THEN answer END) AS pos_business_type,
    MAX(CASE WHEN question = 'gp_fascia' THEN answer END) AS gp_fascia,
    MAX(CASE WHEN question = 'old_posm_counts' THEN answer END) AS old_posm_counts,    
    MAX(CASE WHEN question = 'gp_fascia_type' THEN answer END) AS gp_fascia_type,
    MAX(CASE WHEN question = 'other_fascia' THEN answer END) AS other_fascia,
    MAX(CASE WHEN question = 'which_other_fascia' THEN answer END) AS which_other_fascia,
    -- -- add new -- --
    MAX(CASE WHEN question = 'theamatic' THEN answer END) AS theamatic,
    --MAX(CASE WHEN question = 'themetic_puzzle_block' THEN answer END) AS themetic_puzzle_block,
    MAX(CASE WHEN question = 'business_other_operator' THEN answer END) AS business_other_operator,
    MAX(CASE WHEN question = 'poster_counts' THEN answer END) AS poster_counts,
    --MAX(CASE WHEN question = '10x20_poster_counts' THEN answer END) AS "10x20_poster_counts",
    MAX(CASE WHEN question = 'sim_sell' THEN answer END) AS sim_sell,
    MAX(CASE WHEN question = 'gp_sim_sell' THEN answer END) AS gp_sim_sell,
    MAX(CASE WHEN question = 'other_sim_sell' THEN answer END) AS other_sim_sell,
    MAX(CASE WHEN question = 'finger_print_scanner' THEN answer END) AS finger_print_scanner,
    MAX(CASE WHEN question = 'other_company_scanner' THEN answer END) AS other_company_scanner
FROM 
    gp.market_visit_contact_flow_maps mvcf
WHERE 
    mvcf.visit_date = '{custom_date}'
    AND mvcf.question NOT IN ('confirmation', 'operational_status', 'audio')
GROUP BY 
    visit_id
ORDER BY 
    visit_id
),
posm_data AS (
SELECT
    mvp.visit_id,
        -- --  poster -- 
    --    SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_L_GPfi_DVC_Dec_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_L_GPfi_DVC_Dec_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Offer_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_1 No. Offer_DATA_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Plan_GA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_1 No. Plan_GA_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Bondho Sim_CB_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bondho Sim_CB_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Clear Cut_RC_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Clear Cut_RC_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Limitless_DATA_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_MyGP App_MYGP_NOV_24",
  --      SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_106 Play Pack_PLP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_106 Play Pack_PLP_NOV_24",
       --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_49 & 29 Card_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_49 & 29 Card_DATA_NOV_24",
      --  SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_Toofan_PLP_NOV_24' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_S_Toofan_PLP_NOV_24",
        -- MAR_25 --
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Plan_GA_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_1 No. Plan_GA_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Bondho Sim_CB_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Bondho Sim_CB_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Limitless_DATA_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Combo Offer_BNDL_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Combo Offer_BNDL_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_MyGP App_MYGP_MAR_25",
        -- -- Festoon -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'FST_Sim Bikroy_GA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Sim Bikroy_GA_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'FST_Internet POS_MYGP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Internet POS_MYGP_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'FST_Flexi_RLD_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_NOV_24",
        -- -- Shopscreen -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_Limitless_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_Limitless_DATA_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_ 1 No. Internet Network_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_ 1 No. Internet Network_NET_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'SS_4G Upgrade_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_4G Upgrade_NET_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_Apnar Elakay_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_Apnar Elakay_NET_NOV_24",    
        -- -- Cover Sticker --
        SUM(COALESCE (CASE WHEN posm."name" = 'COVS_1 No. Network_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "COVS_1 No. Network_NET_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'COVS_1 No. Network Internet_ DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "COVS_1 No. Network Internet_ DATA_NOV_24",
        -- -- All others -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_4_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_4_Poster Display Board_BRAND_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_3_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_3_Poster Display Board_BRAND_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_2_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_2_Poster Display Board_BRAND_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'DLR_GPfi_DVC_Dec_24' THEN mvp.amount ELSE 0 END,0)) "DLR_GPfi_DVC_Dec_24" ,
        -- -- New Add -- -- 
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (219)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (219)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (618-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (618-629)_DATA_APR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (818-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (818-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (518-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (518-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (818-728)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (818-728)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (1099-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (1099-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (518-529)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (518-529)_DATA_APR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (198)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (198)_DATA_APR_25",
		-- --
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_MB_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "SS_MB_SKITTO_MAR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Deno_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Deno_SKITTO_MAR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_M_RC_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_M_RC_SKITTO_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Telco_GA_MAR_25' THEN mvp.amount ELSE 0 END,0)) "FST_Telco_GA_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'COVS_Net_1_No._Network_DATA_MAR_25' THEN mvp.amount ELSE 0 END,0)) "COVS_Net_1_No._Network_DATA_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_Mar_25' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_Mar_25",
		-- --
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Flexi_BRND_Feb_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Flexi_BRND_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_GPfi_DVC_Feb_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_GPfi_DVC_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_4G Upgrade_NET_Feb_25' THEN mvp.amount ELSE 0 END,0)) "SS_4G Upgrade_NET_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_MyGP_App_ MYGP_MAR-25' THEN mvp.amount ELSE 0 END,0)) "SS_MyGP_App_ MYGP_MAR-25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "SS_1_No_Network_NET_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "FST_1_No_Network_NET_MAY_25"	,
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_Net_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "SS_Net_1_No_Network_NET_MAY_25",	
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_39_Data_&_29_Voice_SC_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_39_Data_&_29_Voice_SC_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Flexiplan_MYGP_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_Flexiplan_MYGP_MAY_25",
		-- --
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Rate_Cutter_RC_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Rate_Cutter_RC_JUN_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_S_GA_198_Offer_SKITTO_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_GA_198_Offer_SKITTO_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Haat_Bazar_NET_JUN_25' THEN mvp.amount ELSE 0 END,0)) "FST_Haat_Bazar_NET_JUN_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25",
		-- --
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Offer_BNDL_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Offer_BNDL_AUG_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_MyGP App_MYGP_AUG_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Limitless_DATA_AUG_25",
        -- --
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_699_Bundle_Offer_BNDL_AUG 25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_699_Bundle_Offer_BNDL_AUG 25",
        -- --
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_SIM_&_Reload_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "FST_SIM_&_Reload_SKITTO_OCT_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Xtra_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "SS_Xtra_SKITTO_OCT_25",
        -- --
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bioscope+_CNT_Nov_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bioscope+_CNT_Nov_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'COVS_No._1_in_Internet_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "COVS_No._1_in_Internet_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Data_Hero_599_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Data_Hero_599_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Data_Hero_698_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Data_Hero_698_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Hero_699_BNDL_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Hero_699_BNDL_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Hero_999_BNDL_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Hero_999_BNDL_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_No.1_in_Internet_DATA_DEC_25' THEN mvp.amount ELSE 0 END,0)) "SS_No.1_in_Internet_DATA_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_49_Min_&_Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_49_Min_&_Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'ST_49_Min_&_Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "ST_49_Min_&_Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_39 Min_&_49 Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_39 Min_&_49 Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_DEC_25' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_DEC_25",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Data_698_DATA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Data_698_DATA_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Data_499_DATA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Data_499_DATA_JAN_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Bundle_999_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Bundle_999_BNDL_JAN_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Bundle_498_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Bundle_498_BNDL_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_498_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_498_BNDL_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_BNDL_FEB_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_BNDL_FEB_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Monthly_Bundle_BNDL_FEB_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Monthly_Bundle_BNDL_FEB_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Ramadan_Poster__18th Feb_26_10X20' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Ramadan_Poster__18th Feb_26_10X20",
        --SUM(COALESCE(CASE WHEN posm."name" = 'Ebadah_Sticker_9x6' THEN mvp.amount ELSE 0 END,0)) "Ebadah_Sticker_9x6",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Internet_e_1_no_Mar_26' THEN mvp.amount ELSE 0 END,0)) "SS_Internet_e_1_no_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Bondho_Simer_1_no_offer_Mar_26' THEN mvp.amount ELSE 0 END,0)) "FST_Bondho_Simer_1_no_offer_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'BNT_9”x11.5_Mar_26' THEN mvp.amount ELSE 0 END,0)) "BNT_9”x11.5_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_Mar_26' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Hajj_Roaming_Apr_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Hajj_Roaming_Apr_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_29 Min_&_49_Data_SC_Apr_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_29 Min_&_49_Data_SC_Apr_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'COVS_Net_1_No._Network_DATA_APR_26' THEN mvp.amount ELSE 0 END,0)) "COVS_Net_1_No._Network_DATA_APR_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Voice_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Voice_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Monthly_Bundle_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Monthly_Bundle_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_GPFI_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_GPFI_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_GPFI_Order_May_26' THEN mvp.amount ELSE 0 END,0)) "FST_GPFI_Order_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_GPFI_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_GPFI_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Data_2_Deno_219_tk_&_599_tk_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Data_2_Deno_219_tk_&_599_tk_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Internet_e_1_no_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_Internet_e_1_no_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Skitto_Recharge_June_26' THEN mvp.amount ELSE 0 END,0)) "FST_Skitto_Recharge_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_RC_1p_sec_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_RC_1p_sec_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_L700_MHz_metro_&_urban_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_L700_MHz_metro_&_urban_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_L700_MHz_semi_urban_&_rural_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_L700_MHz_semi_urban_&_rural_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_L700MHz_June_26' THEN mvp.amount ELSE 0 END,0)) "SS_L700MHz_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26' THEN mvp.amount ELSE 0 END,0)) "SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Voice_&_Data_Combo_Scratch_Card_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Voice_&_Data_Combo_Scratch_Card_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Power_Load_July_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Power_Load_July_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Churn_Back_July_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Churn_Back_July_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Skitto_Sim_Reload_August_26' THEN mvp.amount ELSE 0 END,0)) "FST_Skitto_Sim_Reload_August_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_August_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_August_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Roaming_On_Aug_26' THEN mvp.amount ELSE 0 END,0)) "SS_Roaming_On_Aug_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Roaming_On_Aug_26' THEN mvp.amount ELSE 0 END,0)) "FST_Roaming_On_Aug_26"
FROM
    gp.market_visit_posm_counts mvp
JOIN
    gp.materials posm ON mvp.posm_id = posm.id
WHERE
    mvp.visit_date = '{custom_date}'
GROUP BY
    mvp.visit_id
ORDER BY
    mvp.visit_id
)
SELECT
    mv.visit_id,
mv.visit_date,
ud. "Circle",
ud. "Region",
ud. "Cluster",
ud. "Territory",
ud. "Distribution House",
ud. "DH Code",
ud. "DH Lat-Long",
mv.route AS "Route",
mv.outlet_code AS "PoS Code",
mv.geo_checked_by AS "Geo Checked By",
mv.is_skitto AS "Is Skitto PoS",
ud. "ME Code", ud. "ME Name",
ud. "ME Contact No.",
ud. "ME Email",
mv.visit_lat_long AS "Visit Geo Location",
mv.operational_status AS "Visit Status",
sd.which_is_right_for_pos as "which_is_right_for_pos",
sd.pos_status as "Pos_Status",
mv.scheduled_visit AS "Scheduled Visit", 
case    
when sd.pre_execution_photo is null or  sd.pre_execution_photo = '' then ''
else concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.pre_execution_photo) 
end AS "Pre Execution",
sd.posm_exists AS "POSM Existance",
sd. "temporarily_&_permanently_closed_photo" AS " Temporary Or Permanently Closed",
sd. "not_found_&_moved_photo" AS "Not Found Or Moved",
sd.old_posm_list as"Old POSM List",
sd.do_posm_remove as "Do POSM Remove",
sd.old_posm_remove_list as"Old POSM Remove List", 
sd.old_posm_counts as "Old POSM Counts",
sd.removed_posm_counts as "Removed POSM Counts",
sd.sim_sell as "Sim_sell",
sd.gp_sim_sell as "Gp_sim_sell",
sd.other_sim_sell as "Other_sim_sell",
sd.finger_print_scanner as "Finger_print_scanner",
sd.other_company_scanner as "Other_company_scanner",
--pd."PSTR_L_GPfi_DVC_Dec_24",
--  pd."PSTR_R_1 No. Offer_DATA_NOV_24",
--  pd."PSTR_R_1 No. Plan_GA_NOV_24",
--pd."PSTR_R_Bondho Sim_CB_NOV_24",
--  pd."PSTR_R_Clear Cut_RC_NOV_24",
--  pd."PSTR_R_Limitless_DATA_NOV_24",
--  pd."PSTR_R_MyGP App_MYGP_NOV_24",
--pd."PSTR_S_106 Play Pack_PLP_NOV_24",
--pd."PSTR_S_49 & 29 Card_DATA_NOV_24",
--pd."PSTR_S_Toofan_PLP_NOV_24",
pd."FST_Sim Bikroy_GA_NOV_24",
--pd."FST_Internet POS_MYGP_NOV_24",
--pd."FST_Flexi_RLD_NOV_24",
pd."SS_Limitless_DATA_NOV_24",
pd."SS_ 1 No. Internet Network_NET_NOV_24",
--pd."SS_4G Upgrade_NET_NOV_24",
pd."SS_Apnar Elakay_NET_NOV_24",
pd."COVS_1 No. Network_NET_NOV_24",
pd."COVS_1 No. Network Internet_ DATA_NOV_24",
pd."PDB_4_Poster Display Board_BRAND_NOV_24",
pd."PDB_3_Poster Display Board_BRAND_NOV_24",
pd."PDB_2_Poster Display Board_BRAND_NOV_24",
--pd."DLR_GPfi_DVC_Dec_24",
--  pd."PSTR_R_1 No. Plan_GA_MAR_25",
--  pd."PSTR_R_Limitless_DATA_MAR_25",
--pd."PSTR_R_Bondho Sim_CB_MAR_25",
--pd."PSTR_R_Combo Offer_BNDL_MAR_25",
--pd."PSTR_R_MyGP App_MYGP_MAR_25",
pd."PSTR_R_Hero Data (219)_DATA_APR_25",
--pd."PSTR_R_Hero Data (618-629)_DATA_APR_25",
--  pd."PSTR_R_Hero Data (818-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (518-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (818-728)_DATA_APR_25",
--pd."PSTR_R_Hero Data (1099-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (518-529)_DATA_APR_25",
--pd."PSTR_R_Hero Data (198)_DATA_APR_25",
pd."SS_MB_SKITTO_MAR_25",
--  pd."PSTR_R_Hero Deno_SKITTO_MAR_25",
--  pd."PSTR_M_RC_SKITTO_MAR_25",
pd."FST_Telco_GA_MAR_25",
pd."COVS_Net_1_No._Network_DATA_MAR_25",
pd."FST_Flexi_RLD_Mar_25",
--pd."PSTR_R_Flexi_BRND_Feb_25",
pd."PSTR_R_GPfi_DVC_Feb_25",
pd."SS_4G Upgrade_NET_Feb_25",
pd."SS_MyGP_App_ MYGP_MAR-25",
pd."SS_1_No_Network_NET_MAY_25",
pd."FST_1_No_Network_NET_MAY_25",
pd."SS_Net_1_No_Network_NET_MAY_25",
--  pd."PSTR_39_Data_&_29_Voice_SC_MAY_25",
pd."PSTR_Flexiplan_MYGP_MAY_25",
pd."PSTR_R_Rate_Cutter_RC_JUN_25",
pd."PSTR_S_GA_198_Offer_SKITTO_MAY_25",
pd."FST_Haat_Bazar_NET_JUN_25",
--  pd."PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25",
--  pd."PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25",
-- pd."PSTR_R_Bundle_Offer_BNDL_AUG_25",
-- pd."PSTR_R_MyGP App_MYGP_AUG_25",
-- pd."PSTR_R_Limitless_DATA_AUG_25",
pd."PSTR_R_699_Bundle_Offer_BNDL_AUG 25",
-- pd."PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25",
pd."FST_SIM_&_Reload_SKITTO_OCT_25",
pd."SS_Xtra_SKITTO_OCT_25",
pd."PSTR_R_Bioscope+_CNT_Nov_25",
pd."COVS_No._1_in_Internet_DATA_NOV_25",
pd."PSTR_R_Data_Hero_599_DATA_NOV_25",
pd."PSTR_R_Data_Hero_698_DATA_NOV_25",
pd."PSTR_R_Bundle_Hero_699_BNDL_NOV_25",
pd."PSTR_R_Bundle_Hero_999_BNDL_NOV_25",
pd."SS_No.1_in_Internet_DATA_DEC_25",
pd."PSTR_49_Min_&_Data_SC_DEC_25",
pd."ST_49_Min_&_Data_SC_DEC_25",
pd."PSTR_39 Min_&_49 Data_SC_DEC_25",
pd."FST_Flexi_RLD_DEC_25",
--pd."PSTR_R_Hero_Data_698_DATA_JAN_26",
pd."PSTR_R_Hero_Data_499_DATA_JAN_26",
--pd."PSTR_R_Hero_Bundle_999_BNDL_JAN_26",
--pd."PSTR_R_Hero_Bundle_498_BNDL_JAN_26",
pd."PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26",
pd."SS_Monthly_Bundle_498_BNDL_JAN_26",
pd."PSTR_R_New_GA_Offer_SKITTO_JAN_26",
pd."SS_Monthly_Bundle_BNDL_FEB_26",
--pd."PSTR_Monthly_Bundle_BNDL_FEB_26",
-- pd."PSTR_R_Ramadan_Poster__18th Feb_26_10X20",
-- pd."Ebadah_Sticker_9x6",
pd."SS_Internet_e_1_no_Mar_26",
pd."FST_Bondho_Simer_1_no_offer_Mar_26",
pd."BNT_9”x11.5_Mar_26",
pd."FST_Flexi_RLD_Mar_26",
pd."PSTR_Hajj_Roaming_Apr_26",
pd."PSTR_29 Min_&_49_Data_SC_Apr_26",
pd."COVS_Net_1_No._Network_DATA_APR_26",
pd."PSTR_Voice_May_26",
pd."PSTR_Monthly_Bundle_May_26",
pd."SS_GPFI_May_26",
pd."FST_GPFI_Order_May_26",
pd."SS_Monthly_Bundle_May_26",
pd."PSTR_GPFI_May_26",
pd."PSTR_Data_2_Deno_219_tk_&_599_tk_May_26",
pd."SS_Internet_e_1_no_May_26",
pd."PSTR_R_New_GA_Offer_SKITTO_June_26",
pd."FST_Skitto_Recharge_June_26",
pd."PSTR_RC_1p_sec_June_26",
pd."PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26",
pd."PSTR_L700_MHz_metro_&_urban_June_26",
pd."PSTR_L700_MHz_semi_urban_&_rural_June_26",
pd."SS_L700MHz_June_26",
pd."SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26",
pd."PSTR_Voice_&_Data_Combo_Scratch_Card_June_26",
pd."PSTR_Power_Load_July_26",
pd."PSTR_Churn_Back_July_26",
pd."FST_Skitto_Sim_Reload_August_26",
pd."PSTR_R_New_GA_Offer_SKITTO_August_26",
pd."SS_Roaming_On_Aug_26",
pd."FST_Roaming_On_Aug_26",
case    
when sd.execution_photo_center is null or  sd.execution_photo_center = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_center) 
end AS "Execution Center",
case    
when sd.execution_photo_left is null or  sd.execution_photo_left = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_left) 
end AS "Execution Left",
case    
when sd.execution_photo_right is null or  sd.execution_photo_right = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_right) 
end AS "Execution Right",
sd.new_posm_counts as "New POSM Counts",
case    
when sd.post_execution_photo is null or  sd.post_execution_photo = '' then ''
else concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/', sd.post_execution_photo) 
end AS "Post Execution",
sd.new_posm_selection as "New POSM Selection",
sd.pos_business_type as "Pos Business Type",
sd.pos_structure as "POS Structure",
sd.pos_location as "POS Place",
sd.gp_fascia as "GP Fascia",
sd.gp_fascia_type AS "GP Fascia Type",
sd.other_fascia AS "Other Fascia",
sd.which_other_fascia AS "Other Fascia List",
sd.business_other_operator as "business_other_operator",
--sd.theamatic as "theamatic",
--sd.themetic_puzzle_block as "themetic_puzzle_block",
sd.poster_counts as "poster_counts"
--sd."10x20_poster_counts" as "10x20_poster_counts"
FROM
    mv_data mv
    LEFT JOIN user_data ud ON mv.me = ud."ME Code"
    LEFT JOIN survey_data sd ON mv.visit_id = sd.visit_id
    LEFT JOIN posm_data pd ON mv.visit_id = pd.visit_id
;
""")

def fetch_raw_data(custom_date, query_function):
    try:
        query = query_function(custom_date)
        df = pd.read_sql_query(query, con=engine)  
        return df
    except Exception as e:
        print(f"❌ Error fetching data: {e}")
        return pd.DataFrame()

def get_day_with_suffix(day):
    if 11 <= day <= 13:
        return f"{day}th"
    last_digit = day % 10
    if last_digit == 1:
        return f"{day}st"
    elif last_digit == 2:
        return f"{day}nd"
    elif last_digit == 3:
        return f"{day}rd"
    else:
        return f"{day}th"

def format_date_for_file_name(start_date_str, end_date_str):
    start = datetime.strptime(start_date_str, "%Y-%m-%d")
    end = datetime.strptime(end_date_str, "%Y-%m-%d")
    start_day = get_day_with_suffix(start.day)
    start_month = start.strftime("%B")
    end_day = get_day_with_suffix(end.day)
    end_month = end.strftime("%B")
    return f"{start_day} {start_month} to {end_day} {end_month}"

def format_sheet_name(custom_date):
    date_obj = datetime.strptime(custom_date, '%Y-%m-%d')
    day = date_obj.day
    if 4 <= day <= 20 or 24 <= day <= 30:
        suffix = "th"
    else:
        suffix = ["st", "nd", "rd"][day % 10 - 1]
    month_name = date_obj.strftime('%B')
    year = date_obj.year
    return f"{day}{suffix} {month_name} -{year}"

def sheet_exists(file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        return sheet_name in wb.sheetnames
    except Exception as e:
        print(f"❌ Error checking sheet existence: {e}")
        return False

def format_excel_file(excel_file_path, sheet_name):
    try:
        wb = load_workbook(excel_file_path)
        ws = wb[sheet_name]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        wb.save(excel_file_path)
        print(f"✅ Excel sheet '{sheet_name}' formatted successfully")
    except Exception as e:
        print(f"❌ Error formatting Excel file: {e}")

def get_or_rename_folder(base_path, start_date, end_date):
    new_range = format_date_for_file_name(start_date, end_date)
    new_folder = f"POS Daily Report ({new_range})"
    new_path = os.path.join(base_path, new_folder)

    for folder in os.listdir(base_path):
        if folder.startswith("POS Daily Report") and os.path.isdir(os.path.join(base_path, folder)):
           
            old_path = os.path.join(base_path, folder)
            if new_folder != folder:
                try:
                    os.rename(old_path, new_path)
                    print(f"♻️ Renamed folder: {folder} ➞ {new_folder}")
                    return new_path
                except Exception as e:
                    print(f"❌ Error renaming folder: {e}")
                    return old_path
            else:
                return old_path

    
    os.makedirs(new_path, exist_ok=True)
    print(f"✅ Created new folder: {new_folder}")
    return new_path

def rename_existing_file_if_needed(folder_path, circle, to_date, custom_date):
    # Allow matching date suffixes like '25th', '3rd', etc.
    pattern = re.compile(
    rf"^{re.escape(circle)}_POS Daily Report \(\d{{1,2}}(?:st|nd|rd|th)? \w+ to \d{{1,2}}(?:st|nd|rd|th)? \w+\)\.xlsx$"
    )

    new_range = format_date_for_file_name(to_date, custom_date)
    new_file_name = f"{circle}_POS Daily Report ({new_range}).xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)

    for fname in os.listdir(folder_path):
        if pattern.match(fname):
            current_file_path = os.path.join(folder_path, fname)

            if fname != new_file_name:
                try:
                    os.rename(current_file_path, new_file_path)
                    print(f"♻️ Renamed file: {fname} ➞ {new_file_name}")
                    return new_file_path
                except Exception as e:
                    print(f"❌ Error renaming file: {e}")
                    return current_file_path
            else:
                print(f"ℹ️ File already named correctly: {fname}")
                return current_file_path

    print(f"📁 No existing file found for {circle}, will create new: {new_file_name}")
    return new_file_path

def zip_files(folder_path, zip_output_path):
    try:
        with zipfile.ZipFile(zip_output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, folder_path)
                    zipf.write(file_path, arcname)
        print(f"✅ Folder zipped at {zip_output_path}")
    except Exception as e:
        print(f"❌ Error zipping files: {e}")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
FROM_EMAIL = "shadman.sayeid@v2.ltd"
EMAIL_PASSWORD = "ujmt gpgi ctbx dohe"
GROUP_MAIL = "cockpit.glm@v2.ltd"
BASE_FOLDER = r"E:/GP/Report/POS Daily Report/Daily POS Visit Reports"
MAX_GMAIL_ATTACHMENT_MB = 24

# Set to True to simulate the Send phase without uploading to Drive or sending email.
DRY_RUN = False


def is_file_locked(file_path):
    """True if the file is open/locked by another program (e.g. Excel)."""
    if not os.path.exists(file_path):
        return False
    try:
        with open(file_path, "r+b"):
            pass
        return False
    except (PermissionError, OSError):
        return True


def save_the_data(to_date, custom_date, base_folder, log=print, progress=None):
    """Generate one formatted .xlsx per circle inside a date-range folder.
    Returns (folder_path, circle_files) where circle_files = {circle: file_path}.

    - Skips circles whose report-date sheet already exists (no duplicate writes).
    - Renames stale same-prefix files to the current date range instead of
      creating conflicting files.
    - Never writes to a file that is currently open in Excel.
    - Verifies each written file (sheet present + row count matches).
    """
    circle_files = {}

    try:
        os.makedirs(base_folder, exist_ok=True)
    except Exception as e:
        log(f"Could not create base folder '{base_folder}': {e}")
        return None, circle_files

    if progress:
        progress(5, "Fetching data from database...")
    df = fetch_raw_data(custom_date, pos_daily_rpt)
    if df.empty:
        log("No data fetched from the database.")
        return None, circle_files

    folder_path = get_or_rename_folder(base_folder, to_date, custom_date)
    log(f"Report folder: {folder_path}")

    circles = df["Circle"].unique()
    if progress:
        progress(10, f"Writing {len(circles)} circle file(s)...")

    for i, circle in enumerate(circles):
        circle_df = df[df["Circle"] == circle]
        file_path = rename_existing_file_if_needed(folder_path, circle, to_date, custom_date)
        if not file_path:
            log(f"Could not determine file path for circle: {circle}")
            continue

        sheet_name = format_sheet_name(custom_date)

        if os.path.exists(file_path) and sheet_exists(file_path, sheet_name):
            log(f"Already downloaded: {os.path.basename(file_path)} (sheet '{sheet_name}' exists). Reusing.")
            circle_files[circle] = file_path
            _remove_stray_sheet(file_path, log)
        else:
            if is_file_locked(file_path):
                log(f"SKIPPED {circle}: {os.path.basename(file_path)} is open in Excel. Close it and re-run.")
            else:
                try:
                    if os.path.exists(file_path):
                        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
                            circle_df.to_excel(writer, index=False, sheet_name=sheet_name)
                    else:
                        with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                            circle_df.to_excel(writer, index=False, sheet_name=sheet_name)
                    _remove_stray_sheet(file_path, log)
                    format_excel_file(file_path, sheet_name)
                    circle_files[circle] = file_path
                    log(f"Downloaded: {os.path.basename(file_path)} ({len(circle_df)} rows)")
                    _verify_file(file_path, sheet_name, len(circle_df), log)
                except Exception as e:
                    log(f"Error writing report for {circle}: {e}. If the file is open in Excel, close it and re-run.")

        if progress:
            progress(10 + int((i + 1) / len(circles) * 60), f"Circle files {i + 1}/{len(circles)}")

    return folder_path, circle_files


def _verify_file(file_path, sheet_name, expected_rows, log=print):
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        actual_rows = max(ws.max_row - 1, 0)
        if actual_rows == expected_rows:
            log(f"Verified {os.path.basename(file_path)}: {actual_rows} rows OK")
        else:
            log(f"WARNING {os.path.basename(file_path)}: {actual_rows} rows, expected {expected_rows}")
    except Exception as e:
        log(f"Verification failed for {os.path.basename(file_path)}: {e}")


def _remove_stray_sheet(file_path, log=print):
    """Remove a leftover empty default 'Sheet' (created by older versions) so the
    file starts with only the real report sheet."""
    try:
        wb = load_workbook(file_path)
        if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
            ws = wb["Sheet"]
            is_empty = all(cell.value is None for row in ws.iter_rows() for cell in row)
            if is_empty:
                wb.remove(ws)
                wb.save(file_path)
                log(f"Removed blank default sheet from {os.path.basename(file_path)}")
    except Exception as e:
        log(f"Could not clean blank sheet in {os.path.basename(file_path)}: {e}")


def _split_emails(value):
    """Split a cell that may contain several emails separated by comma/semicolon/newline."""
    if value is None:
        return []
    parts = re.split(r"[,;\n]+", str(value).strip())
    return [p.strip() for p in parts if p.strip() and p.strip().lower() != "nan"]


def _collect_recipients(headers, rows, log=print):
    """From a header row and a row iterator (tuples), collect To/Cc email lists."""
    lower_headers = [str(h).strip().lower() if h else "" for h in headers]
    to_col = lower_headers.index("to") if "to" in lower_headers else None
    cc_col = lower_headers.index("cc") if "cc" in lower_headers else None

    to_list, cc_list = [], []
    for row in rows:
        if to_col is not None and to_col < len(row) and row[to_col]:
            for val in _split_emails(row[to_col]):
                if val and val not in to_list:
                    to_list.append(val)
        if cc_col is not None and cc_col < len(row) and row[cc_col]:
            for val in _split_emails(row[cc_col]):
                if val and val not in cc_list:
                    cc_list.append(val)

    if to_col is None:
        log("No 'To' column found; only Cc (if any) will be collected.")
    return {"to": to_list, "cc": cc_list}


def _load_recipients_xlsx(xlsx_path, log=print):
    """Excel workbook: one worksheet per Circle, columns 'To' and 'Cc' in row 1."""
    mapping = {}
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        log(f"Could not read recipient Excel file: {e}")
        return mapping

    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]
        rows = ws.iter_rows(min_row=2, values_only=True)
        mapping[ws.title.strip().upper()] = _collect_recipients(headers, rows, log)
        entry = mapping[ws.title.strip().upper()]
        log(f"Recipients for '{ws.title}': {len(entry['to'])} To, {len(entry['cc'])} Cc")

    return mapping


def _load_recipients_xls(xls_path, log=print):
    """Legacy .xls workbook: same one-sheet-per-Circle structure, read via pandas/xlrd."""
    mapping = {}
    try:
        sheets = pd.read_excel(xls_path, sheet_name=None, header=0)
    except Exception as e:
        log(f"Could not read .xls recipient file: {e} (install xlrd or save the file as .xlsx/.csv)")
        return mapping

    for sheet_name, df in sheets.items():
        headers = list(df.columns)
        rows = df.itertuples(index=False, name=None)
        mapping[str(sheet_name).strip().upper()] = _collect_recipients(headers, rows, log)
        entry = mapping[str(sheet_name).strip().upper()]
        log(f"Recipients for '{sheet_name}': {len(entry['to'])} To, {len(entry['cc'])} Cc")

    return mapping


def _load_recipients_csv(csv_path, log=print):
    """CSV recipient file: one row per recipient with columns 'Circle', 'To', 'Cc'.
    A cell may contain several emails separated by comma/semicolon/newline.
    """
    mapping = {}
    try:
        df = pd.read_csv(csv_path, dtype=str)
    except Exception as e:
        log(f"Could not read CSV recipient file: {e}")
        return mapping

    df.columns = [str(c).strip() for c in df.columns]
    lower_cols = {c.lower(): c for c in df.columns}
    circle_col = lower_cols.get("circle")
    to_col = lower_cols.get("to")
    cc_col = lower_cols.get("cc")

    if not circle_col:
        log("CSV recipient file must contain a 'Circle' column (plus optional 'To' and 'Cc').")
        return mapping

    for _, r in df.iterrows():
        circle_val = r[circle_col]
        if circle_val is None or not pd.notna(circle_val) or not str(circle_val).strip():
            continue
        circle = str(circle_val).strip().upper()
        entry = mapping.setdefault(circle, {"to": [], "cc": []})
        if to_col and to_col in df.columns:
            for email in _split_emails(r[to_col]):
                if email and email not in entry["to"]:
                    entry["to"].append(email)
        if cc_col and cc_col in df.columns:
            for email in _split_emails(r[cc_col]):
                if email and email not in entry["cc"]:
                    entry["cc"].append(email)

    for k, v in mapping.items():
        log(f"Recipients for '{k}': {len(v['to'])} To, {len(v['cc'])} Cc")

    return mapping


def load_email_recipients(file_path, log=print):
    """Read a recipient mapping file (Excel or CSV) -> {CIRCLE_UPPER: {"to": [...], "cc": [...]}}.

    - .xlsx / .xls: one worksheet per Circle, columns 'To' and 'Cc' in row 1.
    - .csv: one row per recipient with columns 'Circle', 'To', 'Cc'.
    Every non-empty email is collected and deduplicated.
    """
    ext = os.path.splitext(str(file_path))[1].lower()
    if ext == ".csv":
        return _load_recipients_csv(file_path, log)
    if ext == ".xls":
        return _load_recipients_xls(file_path, log)
    if ext in (".xlsx", ".xlsm"):
        return _load_recipients_xlsx(file_path, log)
    log(f"Unsupported recipient file type: '{ext or '(none)'}'. Use an .xlsx, .xls or .csv file.")
    return {}


def send_circle_email(circle, file_path, to_emails, cc_emails, formatted_range, log=print):
    """Send ONE email for ONE circle: its own report attached + Google Drive link."""
    subject = f"{circle} POS Daily Report ({formatted_range})"

    if not to_emails:
        log(f"Skipping {circle}: no To recipients found.")
        return False
    if not os.path.exists(file_path):
        log(f"Skipping {circle}: file not found at {file_path}")
        return False

    if DRY_RUN:
        log(f"[DRY RUN] Would email {circle} -> To: {', '.join(to_emails)}"
            f"{'; Cc: ' + ', '.join(cc_emails) if cc_emails else ''} | {os.path.basename(file_path)}")
        return True

    try:
        from pydrive2.auth import GoogleAuth
        from pydrive2.drive import GoogleDrive

        gauth = GoogleAuth()
        gauth.LoadClientConfigFile("client_secrets.json")
        gauth.settings["get_refresh_token"] = True
        gauth.settings["oauth_scope"] = ["https://www.googleapis.com/auth/drive"]
        gauth.LoadCredentialsFile("mycreds.txt")
        if gauth.credentials is None:
            gauth.LocalWebserverAuth()
        elif gauth.access_token_expired:
            gauth.Refresh()
        else:
            gauth.Authorize()
        gauth.SaveCredentialsFile("mycreds.txt")

        drive = GoogleDrive(gauth)

        drive_link = None
        try:
            file_drive = drive.CreateFile({"title": os.path.basename(file_path)})
            file_drive.SetContentFile(file_path)
            file_drive.Upload()
            file_drive.InsertPermission({"type": "anyone", "value": "anyone", "role": "reader"})
            drive_link = file_drive["alternateLink"]
        except Exception as e:
            log(f"Drive upload failed for {circle} (continuing with attachment only): {e}")

        disclaimer_html = (
            "<hr>"
            "<small> Disclaimer: The information in this document and attachments is confidential and may also "
            "be legally privileged. It is intended only for the use of the named recipient. Internet communications "
            "are not secure and therefore, Madly Ltd. does not accept legal responsibility for the contents of this "
            "message. If you are not the intended recipient, please notify us immediately and then delete this "
            "document. Do not disclose the contents of this document to any other person, nor take any copies. "
            "Violation of this notice may be unlawful.</small>"
        )
        link_html = f'<p><a href="{drive_link}">{drive_link}</a></p>' if drive_link else ""

        body = (
            "<p>Dear Concern,</p>"
            f"<p>Please find attached the <b>{circle} POS Daily Report ({formatted_range})</b>.</p>"
            f"{link_html}"
            "<p><i>[Please note, This is an automated email report. Please do not reply to this email or mark it "
            "as spam.]</i><br>Regards,<br><b>Cockpit GLM System</b></p>"
            f"{disclaimer_html}"
        )

        msg = MIMEMultipart()
        msg["From"] = GROUP_MAIL
        msg["Subject"] = subject
        msg["To"] = ", ".join(to_emails)
        if cc_emails:
            msg["Cc"] = ", ".join(cc_emails)
        all_recipients = to_emails + cc_emails
        msg.attach(MIMEText(body, "html"))

        max_size = MAX_GMAIL_ATTACHMENT_MB * 1024 * 1024
        attachment_sent = False
        if os.path.getsize(file_path) <= max_size:
            with open(file_path, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(file_path)}"')
            msg.attach(part)
            attachment_sent = True
        else:
            log(f"[{circle}] File too large for Gmail attachment "
                f"({os.path.getsize(file_path) / (1024 * 1024):.1f} MB). Sending Drive link only.")
            body = (
                "<p>Dear Concern,</p>"
                f"<p>Please find attached the <b>{circle} POS Daily Report ({formatted_range})</b>.</p>"
                "<p>The file is too large to attach via Gmail, so it is available only through the Google Drive "
                "link below.</p>"
                f"{link_html}"
                "<p><i>[Please note, This is an automated email report. Please do not reply to this email or mark "
                "it as spam.]</i><br>Regards,<br><b>Cockpit GLM System</b></p>"
                f"{disclaimer_html}"
            )
            msg = MIMEMultipart()
            msg["From"] = GROUP_MAIL
            msg["Subject"] = subject
            msg["To"] = ", ".join(to_emails)
            if cc_emails:
                msg["Cc"] = ", ".join(cc_emails)
            msg.attach(MIMEText(body, "html"))

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(FROM_EMAIL, EMAIL_PASSWORD)
            server.sendmail(GROUP_MAIL, all_recipients, msg.as_string())

        if attachment_sent:
            log(f"[{circle}] Email sent to {', '.join(all_recipients)} (attachment + Drive link)")
        else:
            log(f"[{circle}] Email sent to {', '.join(all_recipients)} (Drive link only)")
        return True
    except Exception as e:
        log(f"[{circle}] Error sending email: {e}")
        return False


class ReportSenderGUI:
    def __init__(self, root):
        self.root = root
        root.title("POS Daily Report - Download & Send")
        root.geometry("740x660")
        root.minsize(660, 580)

        self.email_list_path = ""
        self.busy = False
        self._build_ui()

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        pad = {"padx": 10, "pady": 3}
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="Report Date Range (YYYY-MM-DD):").grid(
            row=0, column=0, columnspan=4, sticky="w", **pad
        )
        ttk.Label(main, text="Start Date:").grid(row=1, column=0, sticky="e", **pad)
        self.start_var = tk.StringVar(value="2026-07-25")
        ttk.Entry(main, textvariable=self.start_var, width=14).grid(row=1, column=1, sticky="w", **pad)
        ttk.Label(main, text="End Date:").grid(row=1, column=2, sticky="e", **pad)
        self.end_var = tk.StringVar(value="2026-07-28")
        ttk.Entry(main, textvariable=self.end_var, width=14).grid(row=1, column=3, sticky="w", **pad)

        ttk.Label(main, text="Recipient mapping file (Excel or CSV):").grid(
            row=2, column=0, columnspan=4, sticky="w", **pad
        )
        self.path_var = tk.StringVar()
        ttk.Entry(main, textvariable=self.path_var, state="readonly").grid(
            row=3, column=0, columnspan=3, sticky="ew", **pad
        )
        ttk.Button(main, text="Browse...", command=self.browse_email_file).grid(
            row=3, column=3, sticky="ew", **pad
        )

        ttk.Label(main, text="Destination folder to save reports:").grid(
            row=4, column=0, columnspan=4, sticky="w", **pad
        )
        self.dest_var = tk.StringVar()
        ttk.Entry(main, textvariable=self.dest_var, state="readonly").grid(
            row=5, column=0, columnspan=3, sticky="ew", **pad
        )
        ttk.Button(main, text="Choose Folder...", command=self.browse_dest_folder).grid(
            row=5, column=3, sticky="ew", **pad
        )

        self.status_var = tk.StringVar(value="Ready. Choose a destination folder (and recipient file for Send).")
        ttk.Label(main, textvariable=self.status_var).grid(row=6, column=0, columnspan=4, sticky="w", **pad)

        self.progress = ttk.Progressbar(main, maximum=100, value=0)
        self.progress.grid(row=7, column=0, columnspan=4, sticky="ew", padx=10, pady=(2, 2))
        self.progress_label = ttk.Label(main, text="0%")
        self.progress_label.grid(row=8, column=0, columnspan=4, sticky="w", padx=10)

        btns = ttk.Frame(main)
        btns.grid(row=9, column=0, columnspan=4, sticky="ew", padx=10, pady=(6, 4))
        self.download_btn = ttk.Button(btns, text="Download Only", command=self.start_download)
        self.download_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.send_btn = ttk.Button(btns, text="Download & Send", command=self.start_download_and_send)
        self.send_btn.pack(side="left", fill="x", expand=True, padx=(4, 0))

        ttk.Label(main, text="Log:").grid(row=10, column=0, columnspan=4, sticky="w", padx=10, pady=(8, 2))
        self.log_box = scrolledtext.ScrolledText(main, height=14, state="disabled", font=("Consolas", 9))
        self.log_box.grid(row=11, column=0, columnspan=4, sticky="nsew", padx=10, pady=(0, 6))

        main.rowconfigure(11, weight=1)
        main.columnconfigure(1, weight=1)
        main.columnconfigure(2, weight=1)
        main.columnconfigure(3, weight=1)

        if os.path.isdir(BASE_FOLDER):
            self.dest_var.set(BASE_FOLDER)

    # ------------------------------------------------------- thread helpers
    def log(self, message):
        def _w():
            self.log_box.config(state="normal")
            self.log_box.insert("end", f"{datetime.now().strftime('%H:%M:%S')}  {message}\n")
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.root.after(0, _w)

    def set_progress(self, value, text=None):
        def _w():
            v = max(0, min(100, int(value)))
            self.progress["value"] = v
            self.progress_label.config(text=f"{v}% - {text}" if text else f"{v}%")
            if text:
                self.status_var.set(text)
        self.root.after(0, _w)

    def set_status(self, text):
        self.root.after(0, lambda: self.status_var.set(text))

    def _set_busy(self):
        self.busy = True
        self.download_btn.config(state="disabled")
        self.send_btn.config(state="disabled")

    def _set_idle(self):
        self.busy = False
        self.download_btn.config(state="normal")
        self.send_btn.config(state="normal")

    # ------------------------------------------------------- file/folder pickers
    def browse_email_file(self):
        path = filedialog.askopenfilename(
            title="Select recipient mapping file (Excel or CSV)",
            filetypes=[("Recipient files", "*.xlsx *.xls *.csv"), ("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv")],
        )
        if not path:
            return
        self.email_list_path = path
        self.path_var.set(path)
        self.log(f"Recipient file selected: {path}")

    def browse_dest_folder(self):
        folder = filedialog.askdirectory(title="Choose destination folder for reports")
        if folder:
            self.dest_var.set(folder)

    # ------------------------------------------------------- validation
    def _validate(self, need_email_list=False):
        start = self.start_var.get().strip()
        end = self.end_var.get().strip()

        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", start) or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", end):
            messagebox.showerror("Invalid Date Format", "Please enter dates in YYYY-MM-DD format.")
            return False
        try:
            datetime.strptime(start, "%Y-%m-%d")
            datetime.strptime(end, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Invalid Date", "The dates entered are not real calendar dates.")
            return False

        dest = self.dest_var.get().strip()
        if not dest:
            messagebox.showerror("Missing folder", "Please choose a destination folder first.")
            return False
        try:
            os.makedirs(dest, exist_ok=True)
        except Exception as e:
            messagebox.showerror("Invalid folder", f"Could not use destination folder: {e}")
            return False

        if need_email_list:
            if not self.email_list_path or not os.path.exists(self.email_list_path):
                messagebox.showerror("Missing file", "Please select a valid recipient mapping Excel file first.")
                return False

        return True

    # ------------------------------------------------------- actions
    def start_download(self):
        if not self._validate():
            return
        self._set_busy()
        self.progress["value"] = 0
        self.set_progress(0, "Starting download...")
        self.log("--- Download Only started ---")
        threading.Thread(target=self._worker_download_only, daemon=True).start()

    def start_download_and_send(self):
        if not self._validate(need_email_list=True):
            return
        self._set_busy()
        self.progress["value"] = 0
        self.set_progress(0, "Starting download & send...")
        self.log("--- Download & Send started ---")
        threading.Thread(target=self._worker_download_and_send, daemon=True).start()

    def _download_phase(self):
        start = self.start_var.get().strip()
        end = self.end_var.get().strip()
        return save_the_data(
            to_date=start,
            custom_date=end,
            base_folder=self.dest_var.get().strip(),
            log=self.log,
            progress=self.set_progress,
        )

    def _zip_phase(self, folder_path):
        if not folder_path or not os.path.isdir(folder_path):
            return
        self.set_progress(72, "Creating ZIP archive...")
        zip_path = folder_path + ".zip"
        zip_files(folder_path, zip_path)
        self.log(f"ZIP saved: {zip_path}")
        self.set_progress(80, "ZIP created.")

    def _worker_download_only(self):
        try:
            folder_path, circle_files = self._download_phase()
            if circle_files:
                self._zip_phase(folder_path)
                self.log(f"Done. {len(circle_files)} circle file(s) in {folder_path}")
                self.set_progress(100, "Download complete.")
            else:
                self.set_progress(0, "No files generated.")
        except Exception as e:
            self.log(f"Unexpected error: {e}")
            self.set_status("Error occurred during download.")
        finally:
            self._set_idle()

    def _worker_download_and_send(self):
        try:
            folder_path, circle_files = self._download_phase()
            if not circle_files:
                self.set_progress(0, "No files generated; nothing to send.")
                self._set_idle()
                return
            self._zip_phase(folder_path)
            self.set_progress(82, "Loading recipients...")
            recipients = load_email_recipients(self.email_list_path, self.log)
            if not recipients:
                self.log("No recipients loaded from the mapping file.")
                self.set_status("Nothing to send (no recipients loaded).")
                self._set_idle()
                return
            self.root.after(0, lambda: self._confirm_and_send(circle_files, recipients))
        except Exception as e:
            self.log(f"Unexpected error: {e}")
            self.set_status("Error occurred.")
            self._set_idle()

    def _confirm_and_send(self, circle_files, recipients):
        sendable = []
        for circle, path in circle_files.items():
            entry = recipients.get(circle.strip().upper(), {})
            if entry.get("to"):
                sendable.append((circle, path, entry.get("to", []), entry.get("cc", [])))

        if not sendable:
            messagebox.showinfo("Nothing to send", "No circles have To recipients in the mapping file.")
            self.log("Nothing to send: no recipient match found.")
            self._finish_send(0, 0, cancelled=True)
            return

        names = ", ".join(c for c, *_ in sendable)
        if not messagebox.askyesno("Confirm send", f"Send emails for {len(sendable)} circle(s):\n{names}?"):
            self.log("Send cancelled by user.")
            self._finish_send(0, 0, cancelled=True)
            return

        self.set_progress(84, "Sending emails...")
        threading.Thread(target=self._worker_send, args=(sendable,), daemon=True).start()

    def _worker_send(self, sendable):
        sent = skipped = 0
        try:
            formatted_range = format_date_for_file_name(
                self.start_var.get().strip(), self.end_var.get().strip()
            )
            for i, (circle, path, to_list, cc_list) in enumerate(sendable):
                ok = send_circle_email(circle, path, to_list, cc_list, formatted_range, self.log)
                sent += 1 if ok else 0
                skipped += 0 if ok else 1
                self.set_progress(84 + int((i + 1) / len(sendable) * 16), f"Sending emails {i + 1}/{len(sendable)}")
        except Exception as e:
            self.log(f"Unexpected error while sending: {e}")
        finally:
            self.root.after(0, lambda: self._finish_send(sent, skipped, cancelled=False))

    def _finish_send(self, sent, skipped, cancelled=False):
        self._set_idle()
        if cancelled:
            self.set_status("Send cancelled.")
            return
        self.log(f"Finished. Emails sent: {sent}, skipped/failed: {skipped}")
        self.set_progress(100, "Complete.")
        self.set_status("Complete. Check the log for details.")


if __name__ == "__main__":
    root = tk.Tk()
    app = ReportSenderGUI(root)
    root.mainloop()
