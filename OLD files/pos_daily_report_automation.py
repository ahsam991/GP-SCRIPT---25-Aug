import os
import shutil
import smtplib
import pandas as pd
import psycopg2
import zipfile
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter, column_index_from_string
from sqlalchemy import create_engine, text
import matplotlib.pyplot as plt
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.utils import make_msgid
from email import encoders
import xlwings as xw
import re
import time
from PIL import ImageGrab
import warnings
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
warnings.filterwarnings("ignore", category=UserWarning, module="xlsxwriter.worksheet")

DB_PARAMS = {
    "dbname": "gp_dev",
    "user": "report_user",
    "password": "report#Gp*User!#__D",
    "host": "gp-stg.cf44ysgum7u8.ap-southeast-1.rds.amazonaws.com",
    "port": 5432  # Adjust if necessary
}

# SQLAlchemy engine
DATABASE_URL = f"postgresql://{DB_PARAMS['user']}:{DB_PARAMS['password']}@{DB_PARAMS['host']}:{DB_PARAMS['port']}/{DB_PARAMS['dbname']}"
engine = create_engine(DATABASE_URL)

def pos_daily_rpt(custom_date):
    return text(f"""
                WITH mv_data AS (
    SELECT
        *
    FROM (
        SELECT
            mvc.id visit_id,
--          mvc.user_id,
            mvc.contacted_by AS me,
            mvc.visit_date,
            r.route_no AS route,
            o.outlet_code,
            o.geo_checked_by,
            o.is_skitto,
            '(' || mvc.lat || ', ' || mvc.long || ')' visit_lat_long,
            mvc.operational_status,
            mvc.scheduled_visit,
            RANK() OVER (PARTITION BY o.outlet_code,mvc.user_id ORDER BY mvc.id DESC)
            rank
        FROM
            gp.market_visit_contacts mvc,
            gp.outlets o,
            gp.routes r
        WHERE
            mvc.location_id = o.id
            AND mvc.location_source_id = r.id
            AND mvc.visit_date = '{custom_date}') x
    WHERE
        rank = 1
        AND x.route not like '%V2TECH%'
),
user_data AS (WITH loc_data AS (
SELECT
    c. "name" AS "Circle",
    r. "name" AS "Region",
    cl. "name" AS "Cluster",
    tr. "name" AS "Territory",
    dh. "name" AS "Distribution House",
    dh.dh_code AS "DH Code",
    '(' || dh.latitude || ', ' || dh.longitude || ')' AS "DH Lat-Long",
    h.id AS dh_id
FROM
    gp.locations c,
    gp.locations r,
    gp.locations cl,
    gp.locations tr,
    gp.locations h,
    gp.house dh
WHERE
    c. "type" = 1
    AND r. "type" = 2
    AND cl. "type" = 3
    AND tr. "type" = 4
    AND h. "type" = 5
    AND r.parent = c.id
    AND cl.parent = r.id
    AND tr.parent = cl.id
    AND h.parent = tr.id
    AND h.is_deleted IS FALSE
    AND h.id = dh.location_id
    AND h. "name" NOT LIKE '%TEST'
)
SELECT
    ld.*,
    u.id AS user_id,
    ui.full_name AS "ME Name",
    ui.official_contact AS "ME Contact No.",
    u.email AS "ME Email",
    b.bundle_code AS "ME Code"
FROM
    gp.users u,
    gp.user_infos ui,
    gp.users_bundle_maps ubm,
    gp.bundles b,
    gp.bundles_house_maps bhm,
    loc_data ld
WHERE
    u.id = ubm.user_id
    AND ui.user_id = u.id
    AND ubm.bundle_id = b.id
    and bhm.bundle_id = b.id
    AND ld.dh_id = bhm.house_id
    AND '{custom_date}' BETWEEN ubm.from_date AND COALESCE(ubm.to_date, '{custom_date}')
      AND '{custom_date}' BETWEEN bhm.from_date AND COALESCE(bhm.to_date, '{custom_date}')
    AND b.bundle_code NOT LIKE '%TEST%'
),
survey_data AS (
SELECT 
    visit_id,
    MAX(CASE WHEN question = 'temporarily_&_permanently_closed_photo' THEN 
        CASE WHEN answer IS NULL OR answer = '' THEN '' 
        ELSE CONCAT('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mvcf.visit_date, 'YYYYMM'),'/', answer) END 
    END) AS "temporarily_&_permanently_closed_photo",
    MAX(CASE WHEN question = 'pre_execution_photo' THEN answer END) AS pre_execution_photo,
    MAX(CASE WHEN question = 'post_execution_photo' THEN answer END) AS post_execution_photo,
    MAX(CASE WHEN question = 'posm_exists' THEN answer END) AS posm_exists,
    MAX(CASE WHEN question = 'not_found_&_moved_photo' THEN 
        CASE WHEN answer IS NULL OR answer = '' THEN '' 
        ELSE CONCAT('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mvcf.visit_date, 'YYYYMM'),'/', answer) END 
    END) AS "not_found_&_moved_photo",
    MAX(CASE WHEN question = 'old_posm_list' THEN answer END) AS old_posm_list,
    MAX(CASE WHEN question = 'do_posm_remove' THEN answer END) AS do_posm_remove,
    MAX(CASE WHEN question = 'old_posm_remove_list' THEN answer END) AS old_posm_remove_list,
    MAX(CASE WHEN question = 'removed_posm_counts' THEN answer END) AS removed_posm_counts,
    MAX(CASE WHEN question = 'new_posm_selection' THEN answer END) AS new_posm_selection,
    MAX(CASE WHEN question = 'new_posm_counts' THEN answer END) AS new_posm_counts, 
--    MAX(CASE WHEN question = 'fascia_type' THEN answer END) AS fascia_type,
--    MAX(CASE WHEN question = 'fascia' THEN answer END) AS fascia,
    MAX(CASE WHEN question = 'execution_photo_right' THEN answer END) AS execution_photo_right,
    MAX(CASE WHEN question = 'execution_photo_left' THEN answer END) AS execution_photo_left,
    MAX(CASE WHEN question = 'execution_photo_center' THEN answer END) AS execution_photo_center,
        MAX(CASE WHEN question = 'which_is_right_for_pos' THEN answer END) AS which_is_right_for_pos,
    MAX(CASE WHEN question = 'pos_status' THEN answer END) AS pos_status,
    MAX(CASE WHEN question = 'pos_structure' THEN answer END) AS pos_structure,
    MAX(CASE WHEN question = 'pos_location' THEN answer END) AS pos_location,
    MAX(CASE WHEN question = 'pos_business_type' THEN answer END) AS pos_business_type,
    MAX(CASE WHEN question = 'gp_fascia' THEN answer END) AS gp_fascia,
    MAX(CASE WHEN question = 'old_posm_counts' THEN answer END) AS old_posm_counts,    
    MAX(CASE WHEN question = 'gp_fascia_type' THEN answer END) AS gp_fascia_type,
    MAX(CASE WHEN question = 'other_fascia' THEN answer END) AS other_fascia,
    MAX(CASE WHEN question = 'which_other_fascia' THEN answer END) AS which_other_fascia,
    -- -- add new -- --
    MAX(CASE WHEN question = 'theamatic' THEN answer END) AS theamatic,
    --MAX(CASE WHEN question = 'themetic_puzzle_block' THEN answer END) AS themetic_puzzle_block,
    MAX(CASE WHEN question = 'business_other_operator' THEN answer END) AS business_other_operator,
    MAX(CASE WHEN question = 'poster_counts' THEN answer END) AS poster_counts,
    --MAX(CASE WHEN question = '10x20_poster_counts' THEN answer END) AS "10x20_poster_counts",
    MAX(CASE WHEN question = 'sim_sell' THEN answer END) AS sim_sell,
    MAX(CASE WHEN question = 'gp_sim_sell' THEN answer END) AS gp_sim_sell,
    MAX(CASE WHEN question = 'other_sim_sell' THEN answer END) AS other_sim_sell,
    MAX(CASE WHEN question = 'finger_print_scanner' THEN answer END) AS finger_print_scanner,
    MAX(CASE WHEN question = 'other_company_scanner' THEN answer END) AS other_company_scanner
FROM 
    gp.market_visit_contact_flow_maps mvcf
WHERE 
    mvcf.visit_date = '{custom_date}'
    AND mvcf.question NOT IN ('confirmation', 'operational_status', 'audio')
GROUP BY 
    visit_id
ORDER BY 
    visit_id
),
posm_data AS (
SELECT
    mvp.visit_id,
        -- --  poster -- 
    --    SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_L_GPfi_DVC_Dec_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_L_GPfi_DVC_Dec_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Offer_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_1 No. Offer_DATA_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Plan_GA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_1 No. Plan_GA_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Bondho Sim_CB_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bondho Sim_CB_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Clear Cut_RC_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Clear Cut_RC_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Limitless_DATA_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_MyGP App_MYGP_NOV_24",
  --      SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_106 Play Pack_PLP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_106 Play Pack_PLP_NOV_24",
       --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_49 & 29 Card_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_49 & 29 Card_DATA_NOV_24",
      --  SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_Toofan_PLP_NOV_24' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_S_Toofan_PLP_NOV_24",
        -- MAR_25 --
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Plan_GA_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_1 No. Plan_GA_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Bondho Sim_CB_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Bondho Sim_CB_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Limitless_DATA_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Combo Offer_BNDL_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Combo Offer_BNDL_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_MyGP App_MYGP_MAR_25",
        -- -- Festoon -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'FST_Sim Bikroy_GA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Sim Bikroy_GA_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'FST_Internet POS_MYGP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Internet POS_MYGP_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'FST_Flexi_RLD_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_NOV_24",
        -- -- Shopscreen -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_Limitless_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_Limitless_DATA_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_ 1 No. Internet Network_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_ 1 No. Internet Network_NET_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'SS_4G Upgrade_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_4G Upgrade_NET_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_Apnar Elakay_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_Apnar Elakay_NET_NOV_24",    
        -- -- Cover Sticker --
        SUM(COALESCE (CASE WHEN posm."name" = 'COVS_1 No. Network_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "COVS_1 No. Network_NET_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'COVS_1 No. Network Internet_ DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "COVS_1 No. Network Internet_ DATA_NOV_24",
        -- -- All others -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_4_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_4_Poster Display Board_BRAND_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_3_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_3_Poster Display Board_BRAND_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_2_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_2_Poster Display Board_BRAND_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'DLR_GPfi_DVC_Dec_24' THEN mvp.amount ELSE 0 END,0)) "DLR_GPfi_DVC_Dec_24" ,
        -- -- New Add -- -- 
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (219)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (219)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (618-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (618-629)_DATA_APR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (818-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (818-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (518-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (518-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (818-728)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (818-728)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (1099-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (1099-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (518-529)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (518-529)_DATA_APR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (198)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (198)_DATA_APR_25",
		-- --
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_MB_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "SS_MB_SKITTO_MAR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Deno_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Deno_SKITTO_MAR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_M_RC_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_M_RC_SKITTO_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Telco_GA_MAR_25' THEN mvp.amount ELSE 0 END,0)) "FST_Telco_GA_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'COVS_Net_1_No._Network_DATA_MAR_25' THEN mvp.amount ELSE 0 END,0)) "COVS_Net_1_No._Network_DATA_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_Mar_25' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_Mar_25",
		-- --
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Flexi_BRND_Feb_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Flexi_BRND_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_GPfi_DVC_Feb_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_GPfi_DVC_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_4G Upgrade_NET_Feb_25' THEN mvp.amount ELSE 0 END,0)) "SS_4G Upgrade_NET_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_MyGP_App_ MYGP_MAR-25' THEN mvp.amount ELSE 0 END,0)) "SS_MyGP_App_ MYGP_MAR-25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "SS_1_No_Network_NET_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "FST_1_No_Network_NET_MAY_25"	,
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_Net_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "SS_Net_1_No_Network_NET_MAY_25",	
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_39_Data_&_29_Voice_SC_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_39_Data_&_29_Voice_SC_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Flexiplan_MYGP_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_Flexiplan_MYGP_MAY_25",
		-- --
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Rate_Cutter_RC_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Rate_Cutter_RC_JUN_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_S_GA_198_Offer_SKITTO_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_GA_198_Offer_SKITTO_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Haat_Bazar_NET_JUN_25' THEN mvp.amount ELSE 0 END,0)) "FST_Haat_Bazar_NET_JUN_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25",
		-- --
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Offer_BNDL_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Offer_BNDL_AUG_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_MyGP App_MYGP_AUG_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Limitless_DATA_AUG_25",
        -- --
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_699_Bundle_Offer_BNDL_AUG 25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_699_Bundle_Offer_BNDL_AUG 25",
        -- --
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_SIM_&_Reload_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "FST_SIM_&_Reload_SKITTO_OCT_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Xtra_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "SS_Xtra_SKITTO_OCT_25",
        -- --
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bioscope+_CNT_Nov_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bioscope+_CNT_Nov_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'COVS_No._1_in_Internet_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "COVS_No._1_in_Internet_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Data_Hero_599_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Data_Hero_599_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Data_Hero_698_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Data_Hero_698_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Hero_699_BNDL_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Hero_699_BNDL_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Hero_999_BNDL_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Hero_999_BNDL_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_No.1_in_Internet_DATA_DEC_25' THEN mvp.amount ELSE 0 END,0)) "SS_No.1_in_Internet_DATA_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_49_Min_&_Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_49_Min_&_Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'ST_49_Min_&_Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "ST_49_Min_&_Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_39 Min_&_49 Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_39 Min_&_49 Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_DEC_25' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_DEC_25",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Data_698_DATA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Data_698_DATA_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Data_499_DATA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Data_499_DATA_JAN_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Bundle_999_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Bundle_999_BNDL_JAN_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Bundle_498_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Bundle_498_BNDL_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_498_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_498_BNDL_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_BNDL_FEB_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_BNDL_FEB_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Monthly_Bundle_BNDL_FEB_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Monthly_Bundle_BNDL_FEB_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Ramadan_Poster__18th Feb_26_10X20' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Ramadan_Poster__18th Feb_26_10X20",
        --SUM(COALESCE(CASE WHEN posm."name" = 'Ebadah_Sticker_9x6' THEN mvp.amount ELSE 0 END,0)) "Ebadah_Sticker_9x6",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Internet_e_1_no_Mar_26' THEN mvp.amount ELSE 0 END,0)) "SS_Internet_e_1_no_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Bondho_Simer_1_no_offer_Mar_26' THEN mvp.amount ELSE 0 END,0)) "FST_Bondho_Simer_1_no_offer_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'BNT_9”x11.5_Mar_26' THEN mvp.amount ELSE 0 END,0)) "BNT_9”x11.5_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_Mar_26' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Hajj_Roaming_Apr_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Hajj_Roaming_Apr_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_29 Min_&_49_Data_SC_Apr_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_29 Min_&_49_Data_SC_Apr_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'COVS_Net_1_No._Network_DATA_APR_26' THEN mvp.amount ELSE 0 END,0)) "COVS_Net_1_No._Network_DATA_APR_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Voice_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Voice_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Monthly_Bundle_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Monthly_Bundle_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_GPFI_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_GPFI_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_GPFI_Order_May_26' THEN mvp.amount ELSE 0 END,0)) "FST_GPFI_Order_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_GPFI_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_GPFI_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Data_2_Deno_219_tk_&_599_tk_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Data_2_Deno_219_tk_&_599_tk_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Internet_e_1_no_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_Internet_e_1_no_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Skitto_Recharge_June_26' THEN mvp.amount ELSE 0 END,0)) "FST_Skitto_Recharge_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_RC_1p_sec_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_RC_1p_sec_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_L700_MHz_metro_&_urban_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_L700_MHz_metro_&_urban_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_L700_MHz_semi_urban_&_rural_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_L700_MHz_semi_urban_&_rural_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_L700MHz_June_26' THEN mvp.amount ELSE 0 END,0)) "SS_L700MHz_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26' THEN mvp.amount ELSE 0 END,0)) "SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Voice_&_Data_Combo_Scratch_Card_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Voice_&_Data_Combo_Scratch_Card_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Power_Load_July_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Power_Load_July_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Churn_Back_July_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Churn_Back_July_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Skitto_Sim_Reload_August_26' THEN mvp.amount ELSE 0 END,0)) "FST_Skitto_Sim_Reload_August_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_August_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_August_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Roaming_On_Aug_26' THEN mvp.amount ELSE 0 END,0)) "SS_Roaming_On_Aug_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Roaming_On_Aug_26' THEN mvp.amount ELSE 0 END,0)) "FST_Roaming_On_Aug_26"
FROM
    gp.market_visit_posm_counts mvp
JOIN
    gp.materials posm ON mvp.posm_id = posm.id
WHERE
    mvp.visit_date = '{custom_date}'
GROUP BY
    mvp.visit_id
ORDER BY
    mvp.visit_id
)
SELECT
    mv.visit_id,
mv.visit_date,
ud. "Circle",
ud. "Region",
ud. "Cluster",
ud. "Territory",
ud. "Distribution House",
ud. "DH Code",
ud. "DH Lat-Long",
mv.route AS "Route",
mv.outlet_code AS "PoS Code",
mv.geo_checked_by AS "Geo Checked By",
mv.is_skitto AS "Is Skitto PoS",
ud. "ME Code", ud. "ME Name",
ud. "ME Contact No.",
ud. "ME Email",
mv.visit_lat_long AS "Visit Geo Location",
mv.operational_status AS "Visit Status",
sd.which_is_right_for_pos as "which_is_right_for_pos",
sd.pos_status as "Pos_Status",
mv.scheduled_visit AS "Scheduled Visit", 
case    
when sd.pre_execution_photo is null or  sd.pre_execution_photo = '' then ''
else concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.pre_execution_photo) 
end AS "Pre Execution",
sd.posm_exists AS "POSM Existance",
sd. "temporarily_&_permanently_closed_photo" AS " Temporary Or Permanently Closed",
sd. "not_found_&_moved_photo" AS "Not Found Or Moved",
sd.old_posm_list as"Old POSM List",
sd.do_posm_remove as "Do POSM Remove",
sd.old_posm_remove_list as"Old POSM Remove List", 
sd.old_posm_counts as "Old POSM Counts",
sd.removed_posm_counts as "Removed POSM Counts",
sd.sim_sell as "Sim_sell",
sd.gp_sim_sell as "Gp_sim_sell",
sd.other_sim_sell as "Other_sim_sell",
sd.finger_print_scanner as "Finger_print_scanner",
sd.other_company_scanner as "Other_company_scanner",
--pd."PSTR_L_GPfi_DVC_Dec_24",
--  pd."PSTR_R_1 No. Offer_DATA_NOV_24",
--  pd."PSTR_R_1 No. Plan_GA_NOV_24",
--pd."PSTR_R_Bondho Sim_CB_NOV_24",
--  pd."PSTR_R_Clear Cut_RC_NOV_24",
--  pd."PSTR_R_Limitless_DATA_NOV_24",
--  pd."PSTR_R_MyGP App_MYGP_NOV_24",
--pd."PSTR_S_106 Play Pack_PLP_NOV_24",
--pd."PSTR_S_49 & 29 Card_DATA_NOV_24",
--pd."PSTR_S_Toofan_PLP_NOV_24",
pd."FST_Sim Bikroy_GA_NOV_24",
--pd."FST_Internet POS_MYGP_NOV_24",
--pd."FST_Flexi_RLD_NOV_24",
pd."SS_Limitless_DATA_NOV_24",
pd."SS_ 1 No. Internet Network_NET_NOV_24",
--pd."SS_4G Upgrade_NET_NOV_24",
pd."SS_Apnar Elakay_NET_NOV_24",
pd."COVS_1 No. Network_NET_NOV_24",
pd."COVS_1 No. Network Internet_ DATA_NOV_24",
pd."PDB_4_Poster Display Board_BRAND_NOV_24",
pd."PDB_3_Poster Display Board_BRAND_NOV_24",
pd."PDB_2_Poster Display Board_BRAND_NOV_24",
--pd."DLR_GPfi_DVC_Dec_24",
--  pd."PSTR_R_1 No. Plan_GA_MAR_25",
--  pd."PSTR_R_Limitless_DATA_MAR_25",
--pd."PSTR_R_Bondho Sim_CB_MAR_25",
--pd."PSTR_R_Combo Offer_BNDL_MAR_25",
--pd."PSTR_R_MyGP App_MYGP_MAR_25",
pd."PSTR_R_Hero Data (219)_DATA_APR_25",
--pd."PSTR_R_Hero Data (618-629)_DATA_APR_25",
--  pd."PSTR_R_Hero Data (818-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (518-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (818-728)_DATA_APR_25",
--pd."PSTR_R_Hero Data (1099-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (518-529)_DATA_APR_25",
--pd."PSTR_R_Hero Data (198)_DATA_APR_25",
pd."SS_MB_SKITTO_MAR_25",
--  pd."PSTR_R_Hero Deno_SKITTO_MAR_25",
--  pd."PSTR_M_RC_SKITTO_MAR_25",
pd."FST_Telco_GA_MAR_25",
pd."COVS_Net_1_No._Network_DATA_MAR_25",
pd."FST_Flexi_RLD_Mar_25",
--pd."PSTR_R_Flexi_BRND_Feb_25",
pd."PSTR_R_GPfi_DVC_Feb_25",
pd."SS_4G Upgrade_NET_Feb_25",
pd."SS_MyGP_App_ MYGP_MAR-25",
pd."SS_1_No_Network_NET_MAY_25",
pd."FST_1_No_Network_NET_MAY_25",
pd."SS_Net_1_No_Network_NET_MAY_25",
--  pd."PSTR_39_Data_&_29_Voice_SC_MAY_25",
pd."PSTR_Flexiplan_MYGP_MAY_25",
pd."PSTR_R_Rate_Cutter_RC_JUN_25",
pd."PSTR_S_GA_198_Offer_SKITTO_MAY_25",
pd."FST_Haat_Bazar_NET_JUN_25",
--  pd."PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25",
--  pd."PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25",
-- pd."PSTR_R_Bundle_Offer_BNDL_AUG_25",
-- pd."PSTR_R_MyGP App_MYGP_AUG_25",
-- pd."PSTR_R_Limitless_DATA_AUG_25",
pd."PSTR_R_699_Bundle_Offer_BNDL_AUG 25",
-- pd."PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25",
pd."FST_SIM_&_Reload_SKITTO_OCT_25",
pd."SS_Xtra_SKITTO_OCT_25",
pd."PSTR_R_Bioscope+_CNT_Nov_25",
pd."COVS_No._1_in_Internet_DATA_NOV_25",
pd."PSTR_R_Data_Hero_599_DATA_NOV_25",
pd."PSTR_R_Data_Hero_698_DATA_NOV_25",
pd."PSTR_R_Bundle_Hero_699_BNDL_NOV_25",
pd."PSTR_R_Bundle_Hero_999_BNDL_NOV_25",
pd."SS_No.1_in_Internet_DATA_DEC_25",
pd."PSTR_49_Min_&_Data_SC_DEC_25",
pd."ST_49_Min_&_Data_SC_DEC_25",
pd."PSTR_39 Min_&_49 Data_SC_DEC_25",
pd."FST_Flexi_RLD_DEC_25",
--pd."PSTR_R_Hero_Data_698_DATA_JAN_26",
pd."PSTR_R_Hero_Data_499_DATA_JAN_26",
--pd."PSTR_R_Hero_Bundle_999_BNDL_JAN_26",
--pd."PSTR_R_Hero_Bundle_498_BNDL_JAN_26",
pd."PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26",
pd."SS_Monthly_Bundle_498_BNDL_JAN_26",
pd."PSTR_R_New_GA_Offer_SKITTO_JAN_26",
pd."SS_Monthly_Bundle_BNDL_FEB_26",
--pd."PSTR_Monthly_Bundle_BNDL_FEB_26",
-- pd."PSTR_R_Ramadan_Poster__18th Feb_26_10X20",
-- pd."Ebadah_Sticker_9x6",
pd."SS_Internet_e_1_no_Mar_26",
pd."FST_Bondho_Simer_1_no_offer_Mar_26",
pd."BNT_9”x11.5_Mar_26",
pd."FST_Flexi_RLD_Mar_26",
pd."PSTR_Hajj_Roaming_Apr_26",
pd."PSTR_29 Min_&_49_Data_SC_Apr_26",
pd."COVS_Net_1_No._Network_DATA_APR_26",
pd."PSTR_Voice_May_26",
pd."PSTR_Monthly_Bundle_May_26",
pd."SS_GPFI_May_26",
pd."FST_GPFI_Order_May_26",
pd."SS_Monthly_Bundle_May_26",
pd."PSTR_GPFI_May_26",
pd."PSTR_Data_2_Deno_219_tk_&_599_tk_May_26",
pd."SS_Internet_e_1_no_May_26",
pd."PSTR_R_New_GA_Offer_SKITTO_June_26",
pd."FST_Skitto_Recharge_June_26",
pd."PSTR_RC_1p_sec_June_26",
pd."PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26",
pd."PSTR_L700_MHz_metro_&_urban_June_26",
pd."PSTR_L700_MHz_semi_urban_&_rural_June_26",
pd."SS_L700MHz_June_26",
pd."SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26",
pd."PSTR_Voice_&_Data_Combo_Scratch_Card_June_26",
pd."PSTR_Power_Load_July_26",
pd."PSTR_Churn_Back_July_26",
pd."FST_Skitto_Sim_Reload_August_26",
pd."PSTR_R_New_GA_Offer_SKITTO_August_26",
pd."SS_Roaming_On_Aug_26",
pd."FST_Roaming_On_Aug_26",
case    
when sd.execution_photo_center is null or  sd.execution_photo_center = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_center) 
end AS "Execution Center",
case    
when sd.execution_photo_left is null or  sd.execution_photo_left = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_left) 
end AS "Execution Left",
case    
when sd.execution_photo_right is null or  sd.execution_photo_right = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_right) 
end AS "Execution Right",
sd.new_posm_counts as "New POSM Counts",
case    
when sd.post_execution_photo is null or  sd.post_execution_photo = '' then ''
else concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/', sd.post_execution_photo) 
end AS "Post Execution",
sd.new_posm_selection as "New POSM Selection",
sd.pos_business_type as "Pos Business Type",
sd.pos_structure as "POS Structure",
sd.pos_location as "POS Place",
sd.gp_fascia as "GP Fascia",
sd.gp_fascia_type AS "GP Fascia Type",
sd.other_fascia AS "Other Fascia",
sd.which_other_fascia AS "Other Fascia List",
sd.business_other_operator as "business_other_operator",
--sd.theamatic as "theamatic",
--sd.themetic_puzzle_block as "themetic_puzzle_block",
sd.poster_counts as "poster_counts"
--sd."10x20_poster_counts" as "10x20_poster_counts"
FROM
    mv_data mv
    LEFT JOIN user_data ud ON mv.me = ud."ME Code"
    LEFT JOIN survey_data sd ON mv.visit_id = sd.visit_id
    LEFT JOIN posm_data pd ON mv.visit_id = pd.visit_id
;
""")

def fetch_raw_data(custom_date, query_function):
    try:
        query = query_function(custom_date)
        df = pd.read_sql_query(query, con=engine)  
        return df
    except Exception as e:
        print(f"❌ Error fetching data: {e}")
        return pd.DataFrame()

def get_day_with_suffix(day):
    if 11 <= day <= 13:
        return f"{day}th"
    last_digit = day % 10
    if last_digit == 1:
        return f"{day}st"
    elif last_digit == 2:
        return f"{day}nd"
    elif last_digit == 3:
        return f"{day}rd"
    else:
        return f"{day}th"

def format_date_for_file_name(start_date_str, end_date_str):
    start = datetime.strptime(start_date_str, "%Y-%m-%d")
    end = datetime.strptime(end_date_str, "%Y-%m-%d")
    start_day = get_day_with_suffix(start.day)
    start_month = start.strftime("%B")
    end_day = get_day_with_suffix(end.day)
    end_month = end.strftime("%B")
    return f"{start_day} {start_month} to {end_day} {end_month}"

def format_sheet_name(custom_date):
    date_obj = datetime.strptime(custom_date, '%Y-%m-%d')
    day = date_obj.day
    if 4 <= day <= 20 or 24 <= day <= 30:
        suffix = "th"
    else:
        suffix = ["st", "nd", "rd"][day % 10 - 1]
    month_name = date_obj.strftime('%B')
    year = date_obj.year
    return f"{day}{suffix} {month_name} -{year}"

def sheet_exists(file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        return sheet_name in wb.sheetnames
    except Exception as e:
        print(f"❌ Error checking sheet existence: {e}")
        return False

def format_excel_file(excel_file_path, sheet_name):
    try:
        wb = load_workbook(excel_file_path)
        ws = wb[sheet_name]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        wb.save(excel_file_path)
        print(f"✅ Excel sheet '{sheet_name}' formatted successfully")
    except Exception as e:
        print(f"❌ Error formatting Excel file: {e}")


def get_or_rename_folder(base_path, start_date, end_date):
    new_range = format_date_for_file_name(start_date, end_date)
    new_folder = f"POS Daily Report ({new_range})"
    new_path = os.path.join(base_path, new_folder)

    for folder in os.listdir(base_path):
        if folder.startswith("POS Daily Report") and os.path.isdir(os.path.join(base_path, folder)):
           
            old_path = os.path.join(base_path, folder)
            if new_folder != folder:
                try:
                    os.rename(old_path, new_path)
                    print(f"♻️ Renamed folder: {folder} ➞ {new_folder}")
                    return new_path
                except Exception as e:
                    print(f"❌ Error renaming folder: {e}")
                    return old_path
            else:
                return old_path

    
    os.makedirs(new_path, exist_ok=True)
    print(f"✅ Created new folder: {new_folder}")
    return new_path


def rename_existing_file_if_needed(folder_path, circle, to_date, custom_date):
    # Allow matching date suffixes like '25th', '3rd', etc.
    pattern = re.compile(
    rf"^{re.escape(circle)}_POS Daily Report \(\d{{1,2}}(?:st|nd|rd|th)? \w+ to \d{{1,2}}(?:st|nd|rd|th)? \w+\)\.xlsx$"
    )

    new_range = format_date_for_file_name(to_date, custom_date)
    new_file_name = f"{circle}_POS Daily Report ({new_range}).xlsx"
    new_file_path = os.path.join(folder_path, new_file_name)

    for fname in os.listdir(folder_path):
        if pattern.match(fname):
            current_file_path = os.path.join(folder_path, fname)

            if fname != new_file_name:
                try:
                    os.rename(current_file_path, new_file_path)
                    print(f"♻️ Renamed file: {fname} ➞ {new_file_name}")
                    return new_file_path
                except Exception as e:
                    print(f"❌ Error renaming file: {e}")
                    return current_file_path
            else:
                print(f"ℹ️ File already named correctly: {fname}")
                return current_file_path

    print(f"📁 No existing file found for {circle}, will create new: {new_file_name}")
    return new_file_path

def save_the_data():
    df = fetch_raw_data(custom_date, pos_daily_rpt)
    if df.empty:
        print("❌ No data fetched from the database.")
        return

    base_folder = r'E:/GP/Report/POS Daily Report/Daily POS Visit Reports'
    folder_path = get_or_rename_folder(base_folder, to_date, custom_date)

    for circle in df['Circle'].unique():
        circle_df = df[df['Circle'] == circle]

        file_path = rename_existing_file_if_needed(folder_path, circle, to_date, custom_date)
        if not file_path:
            print(f"❌ Could not determine file path for circle: {circle}")
            continue

        sheet_name = format_sheet_name(custom_date)

        if os.path.exists(file_path) and sheet_exists(file_path, sheet_name):
            print(f"⏩ Sheet '{sheet_name}' already exists in {os.path.basename(file_path)}. Skipping...")
            continue

        if not os.path.exists(file_path):
            wb = Workbook()
            wb.save(file_path)
            print(f"✅ Created new Excel file: {file_path}")
        else:
            print(f"ℹ️ File exists: {file_path}")

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                circle_df.to_excel(writer, index=False, sheet_name=sheet_name)
            format_excel_file(file_path, sheet_name)
        except Exception as e:
            print(f"❌ Error writing to Excel: {e}")


def zip_files(folder_path, zip_output_path):
    try:
        with zipfile.ZipFile(zip_output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, folder_path)
                    zipf.write(file_path, arcname)
        print(f"✅ Folder zipped at {zip_output_path}")
    except Exception as e:
        print(f"❌ Error zipping files: {e}")

def send_email_with_attachments(subject, body, to_emails, cc_emails, attachments):
    from_email = "shadman.sayeid@v2.ltd"
    # email_password = "ujmt gpgi ctbx dohe"
    group_mail = "cockpit.glm@v2.ltd"

    try:
        from pydrive2.auth import GoogleAuth
        from pydrive2.drive import GoogleDrive

        gauth = GoogleAuth()
        gauth.LoadClientConfigFile("client_secrets.json")
        gauth.settings['get_refresh_token'] = True
        gauth.settings['oauth_scope'] = ['https://www.googleapis.com/auth/drive']
        gauth.LoadCredentialsFile("mycreds.txt")
        if gauth.credentials is None:
            gauth.LocalWebserverAuth()
        elif gauth.access_token_expired:
            gauth.Refresh()
        else:
            gauth.Authorize()
        gauth.SaveCredentialsFile("mycreds.txt")

        
        drive = GoogleDrive(gauth)

        drive_links = []
        for attachment_path in attachments:
            if os.path.exists(attachment_path):
                file_drive = drive.CreateFile({'title': os.path.basename(attachment_path)})
                file_drive.SetContentFile(attachment_path)
                file_drive.Upload()
                file_drive.InsertPermission({'type': 'anyone', 'value': 'anyone', 'role': 'reader'})
                drive_links.append(file_drive['alternateLink'])
            else:
                print(f"⚠️ Attachment not found: {attachment_path}")

        # Build HTML body with clickable links and disclaimer in <small>
        disclaimer_html = (
            "<hr>"
            "<small> Disclaimer: The information in this document and attachments is confidential and may also be legally privileged. It is intended only for the use of the named recipient. Internet communications are not secure and therefore, Madly Ltd. does not accept legal responsibility for the contents of this message. If you are not the intended recipient, please notify us immediately and then delete this document. Do not disclose the contents of this document to any other person, nor take any copies. Violation of this notice may be unlawful." 
            "</small>"
        )
        if drive_links:
            links_html = '<br>'.join(f'<a href="{link}">{link}</a>' for link in drive_links)
            body = (
                f"<p>Dear Concern,</p>"
                f"<p>Please find the <b>POS Daily Reports circle wise ({subject.split('(')[-1].rstrip(')')})</b> herewith:</p>"
                f"<p>{links_html}</p>"
                f"<p><i>[Please note, This is an automated email report. Please do not reply to this email or mark it as spam.]</i><br>"
                f"Regards,<br><b>Cockpit GLM System</b></p>"
                f"{disclaimer_html}"
            )
        else:
            body = (
                f"<p>Dear Concern,</p>"
                f"<p>The POS Daily Report is not available for download.</p>"
                f"<p>Please note, This is an automated email report. Please do not reply to this email or mark it as spam.<br>"
                f"Regards,<br>Cockpit GLM System</p>"
                f"{disclaimer_html}"
            )

        msg = MIMEMultipart()
        msg['From'] = group_mail
        msg['Subject'] = subject
        msg['To'] = ", ".join(to_emails)
        msg['Cc'] = ", ".join(cc_emails)
        all_recipients = to_emails + cc_emails
        msg.attach(MIMEText(body, 'html'))

        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(from_email, email_password)
            server.sendmail(group_mail, all_recipients, msg.as_string())
            print(f"✅ Email sent to {', '.join(all_recipients)}")
    except Exception as e:
        print(f"❌ Error sending email: {e}")


if __name__ == "__main__":
    custom_date = '2026-07-28'
    # custom_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')  
    to_date = '2026-07-25'  
    formatted_range = format_date_for_file_name(to_date, custom_date)
    custom_date_filename = datetime.strptime(custom_date, '%Y-%m-%d').strftime('%d %B %Y')

    save_the_data()

    base_folder = r'E:/GP/Report/POS Daily Report/Daily POS Visit Reports'
    folder_name = f"POS Daily Report ({formatted_range})"
    folder_path = os.path.join(base_folder, folder_name)
    zip_path = folder_path + ".zip"
    zip_files(folder_path, zip_path)
    print(f"✅ Excel file copied to: {zip_path}")
    
    # to_emails = ['ssshaivik@ms365ce.onmicrosoft.com']  # Add more emails as needed
    # cc_emails = ['ruhulaminshazid@gmail.com']  # Add CC emails if needed

    # # # Directly set the email recipients here
    to_emails = ['saju@ims.net.bd','rakib@ims.com.bd','dip@asiaticexp.com','obydur.gpme@asiaticexp.com','subrata@asiaticexp.com','shemul.gpme@asiaticexp.com','riadbiswas03@gmail.com','abu.rayhan@grameenphone.com','a_amin@grameenphone.com','nsarker@grameenphone.com']  # Add more emails as needed
    cc_emails = ['mahamudul@grameenphone.com','jubair.ovi@v2.ltd']  # Add CC emails if needed

    
    subject = f"POS Daily Report ({formatted_range})"
    # The body is built inside send_email_with_attachments; no need to pass a template here
    send_email_with_attachments(subject, None, to_emails, cc_emails, [zip_path])
    
    