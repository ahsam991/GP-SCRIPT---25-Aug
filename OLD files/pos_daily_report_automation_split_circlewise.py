import os
import shutil
import smtplib
import pandas as pd
import psycopg2
import zipfile
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter, column_index_from_string
from sqlalchemy import create_engine, text
import matplotlib.pyplot as plt
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.utils import make_msgid
from email import encoders
import xlwings as xw
import re
import time
from PIL import ImageGrab
import warnings
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import threading
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
warnings.filterwarnings("ignore", category=UserWarning, module="xlsxwriter.worksheet")

DB_PARAMS = {
    "dbname": "gp_dev",
    "user": "report_user",
    "password": "report#Gp*User!#__D",
    "host": "gp-stg.cf44ysgum7u8.ap-southeast-1.rds.amazonaws.com",
    "port": 5432  # Adjust if necessary
}

# SQLAlchemy engine
DATABASE_URL = f"postgresql://{DB_PARAMS['user']}:{DB_PARAMS['password']}@{DB_PARAMS['host']}:{DB_PARAMS['port']}/{DB_PARAMS['dbname']}"
engine = create_engine(DATABASE_URL)

def pos_daily_rpt(to_date, custom_date):
    return text(f"""
                WITH mv_data AS (
    SELECT
        *
    FROM (
        SELECT
            mvc.id visit_id,
--          mvc.user_id,
            mvc.contacted_by AS me,
            mvc.visit_date,
            r.route_no AS route,
            o.outlet_code,
            o.geo_checked_by,
            o.is_skitto,
            '(' || mvc.lat || ', ' || mvc.long || ')' visit_lat_long,
            mvc.operational_status,
            mvc.scheduled_visit,
            RANK() OVER (PARTITION BY o.outlet_code,mvc.user_id ORDER BY mvc.id DESC)
            rank
        FROM
            gp.market_visit_contacts mvc,
            gp.outlets o,
            gp.routes r
        WHERE
            mvc.location_id = o.id
            AND mvc.location_source_id = r.id
            AND mvc.visit_date BETWEEN '{to_date}' AND '{custom_date}') x
    WHERE
        rank = 1
        AND x.route not like '%V2TECH%'
),
user_data AS (WITH loc_data AS (
SELECT
    c. "name" AS "Circle",
    r. "name" AS "Region",
    cl. "name" AS "Cluster",
    tr. "name" AS "Territory",
    dh. "name" AS "Distribution House",
    dh.dh_code AS "DH Code",
    '(' || dh.latitude || ', ' || dh.longitude || ')' AS "DH Lat-Long",
    h.id AS dh_id
FROM
    gp.locations c,
    gp.locations r,
    gp.locations cl,
    gp.locations tr,
    gp.locations h,
    gp.house dh
WHERE
    c. "type" = 1
    AND r. "type" = 2
    AND cl. "type" = 3
    AND tr. "type" = 4
    AND h. "type" = 5
    AND r.parent = c.id
    AND cl.parent = r.id
    AND tr.parent = cl.id
    AND h.parent = tr.id
    AND h.is_deleted IS FALSE
    AND h.id = dh.location_id
    AND h. "name" NOT LIKE '%TEST'
)
SELECT
    ld.*,
    u.id AS user_id,
    ui.full_name AS "ME Name",
    ui.official_contact AS "ME Contact No.",
    u.email AS "ME Email",
    b.bundle_code AS "ME Code"
FROM
    gp.users u,
    gp.user_infos ui,
    gp.users_bundle_maps ubm,
    gp.bundles b,
    gp.bundles_house_maps bhm,
    loc_data ld
WHERE
    u.id = ubm.user_id
    AND ui.user_id = u.id
    AND ubm.bundle_id = b.id
    and bhm.bundle_id = b.id
    AND ld.dh_id = bhm.house_id
    AND ubm.from_date <= '{custom_date}' AND COALESCE(ubm.to_date, '{custom_date}') >= '{to_date}'
      AND bhm.from_date <= '{custom_date}' AND COALESCE(bhm.to_date, '{custom_date}') >= '{to_date}'
    AND b.bundle_code NOT LIKE '%TEST%'
),
survey_data AS (
SELECT 
    visit_id,
    MAX(CASE WHEN question = 'temporarily_&_permanently_closed_photo' THEN 
        CASE WHEN answer IS NULL OR answer = '' THEN '' 
        ELSE CONCAT('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mvcf.visit_date, 'YYYYMM'),'/', answer) END 
    END) AS "temporarily_&_permanently_closed_photo",
    MAX(CASE WHEN question = 'pre_execution_photo' THEN answer END) AS pre_execution_photo,
    MAX(CASE WHEN question = 'post_execution_photo' THEN answer END) AS post_execution_photo,
    MAX(CASE WHEN question = 'posm_exists' THEN answer END) AS posm_exists,
    MAX(CASE WHEN question = 'not_found_&_moved_photo' THEN 
        CASE WHEN answer IS NULL OR answer = '' THEN '' 
        ELSE CONCAT('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mvcf.visit_date, 'YYYYMM'),'/', answer) END 
    END) AS "not_found_&_moved_photo",
    MAX(CASE WHEN question = 'old_posm_list' THEN answer END) AS old_posm_list,
    MAX(CASE WHEN question = 'do_posm_remove' THEN answer END) AS do_posm_remove,
    MAX(CASE WHEN question = 'old_posm_remove_list' THEN answer END) AS old_posm_remove_list,
    MAX(CASE WHEN question = 'removed_posm_counts' THEN answer END) AS removed_posm_counts,
    MAX(CASE WHEN question = 'new_posm_selection' THEN answer END) AS new_posm_selection,
    MAX(CASE WHEN question = 'new_posm_counts' THEN answer END) AS new_posm_counts, 
--    MAX(CASE WHEN question = 'fascia_type' THEN answer END) AS fascia_type,
--    MAX(CASE WHEN question = 'fascia' THEN answer END) AS fascia,
    MAX(CASE WHEN question = 'execution_photo_right' THEN answer END) AS execution_photo_right,
    MAX(CASE WHEN question = 'execution_photo_left' THEN answer END) AS execution_photo_left,
    MAX(CASE WHEN question = 'execution_photo_center' THEN answer END) AS execution_photo_center,
        MAX(CASE WHEN question = 'which_is_right_for_pos' THEN answer END) AS which_is_right_for_pos,
    MAX(CASE WHEN question = 'pos_status' THEN answer END) AS pos_status,
    MAX(CASE WHEN question = 'pos_structure' THEN answer END) AS pos_structure,
    MAX(CASE WHEN question = 'pos_location' THEN answer END) AS pos_location,
    MAX(CASE WHEN question = 'pos_business_type' THEN answer END) AS pos_business_type,
    MAX(CASE WHEN question = 'gp_fascia' THEN answer END) AS gp_fascia,
    MAX(CASE WHEN question = 'old_posm_counts' THEN answer END) AS old_posm_counts,    
    MAX(CASE WHEN question = 'gp_fascia_type' THEN answer END) AS gp_fascia_type,
    MAX(CASE WHEN question = 'other_fascia' THEN answer END) AS other_fascia,
    MAX(CASE WHEN question = 'which_other_fascia' THEN answer END) AS which_other_fascia,
    -- -- add new -- --
    MAX(CASE WHEN question = 'theamatic' THEN answer END) AS theamatic,
    --MAX(CASE WHEN question = 'themetic_puzzle_block' THEN answer END) AS themetic_puzzle_block,
    MAX(CASE WHEN question = 'business_other_operator' THEN answer END) AS business_other_operator,
    MAX(CASE WHEN question = 'poster_counts' THEN answer END) AS poster_counts,
    --MAX(CASE WHEN question = '10x20_poster_counts' THEN answer END) AS "10x20_poster_counts",
    MAX(CASE WHEN question = 'sim_sell' THEN answer END) AS sim_sell,
    MAX(CASE WHEN question = 'gp_sim_sell' THEN answer END) AS gp_sim_sell,
    MAX(CASE WHEN question = 'other_sim_sell' THEN answer END) AS other_sim_sell,
    MAX(CASE WHEN question = 'finger_print_scanner' THEN answer END) AS finger_print_scanner,
    MAX(CASE WHEN question = 'other_company_scanner' THEN answer END) AS other_company_scanner
FROM 
    gp.market_visit_contact_flow_maps mvcf
WHERE 
    mvcf.visit_date BETWEEN '{to_date}' AND '{custom_date}'
    AND mvcf.question NOT IN ('confirmation', 'operational_status', 'audio')
GROUP BY 
    visit_id
ORDER BY 
    visit_id
),
posm_data AS (
SELECT
    mvp.visit_id,
        -- --  poster -- 
    --    SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_L_GPfi_DVC_Dec_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_L_GPfi_DVC_Dec_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Offer_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_1 No. Offer_DATA_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Plan_GA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_1 No. Plan_GA_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Bondho Sim_CB_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bondho Sim_CB_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Clear Cut_RC_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Clear Cut_RC_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Limitless_DATA_NOV_24",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_MyGP App_MYGP_NOV_24",
  --      SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_106 Play Pack_PLP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_106 Play Pack_PLP_NOV_24",
       --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_49 & 29 Card_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_49 & 29 Card_DATA_NOV_24",
      --  SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_S_Toofan_PLP_NOV_24' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_S_Toofan_PLP_NOV_24",
        -- MAR_25 --
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_1 No. Plan_GA_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_1 No. Plan_GA_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Bondho Sim_CB_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Bondho Sim_CB_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Limitless_DATA_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_Combo Offer_BNDL_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_Combo Offer_BNDL_MAR_25",
        --SUM(COALESCE (CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_MAR_25' THEN mvp.amount ELSE 0 END,0)) AS "PSTR_R_MyGP App_MYGP_MAR_25",
        -- -- Festoon -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'FST_Sim Bikroy_GA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Sim Bikroy_GA_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'FST_Internet POS_MYGP_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Internet POS_MYGP_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'FST_Flexi_RLD_NOV_24' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_NOV_24",
        -- -- Shopscreen -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_Limitless_DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_Limitless_DATA_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_ 1 No. Internet Network_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_ 1 No. Internet Network_NET_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'SS_4G Upgrade_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_4G Upgrade_NET_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'SS_Apnar Elakay_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "SS_Apnar Elakay_NET_NOV_24",    
        -- -- Cover Sticker --
        SUM(COALESCE (CASE WHEN posm."name" = 'COVS_1 No. Network_NET_NOV_24' THEN mvp.amount ELSE 0 END,0)) "COVS_1 No. Network_NET_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'COVS_1 No. Network Internet_ DATA_NOV_24' THEN mvp.amount ELSE 0 END,0)) "COVS_1 No. Network Internet_ DATA_NOV_24",
        -- -- All others -- 
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_4_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_4_Poster Display Board_BRAND_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_3_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_3_Poster Display Board_BRAND_NOV_24",
        SUM(COALESCE (CASE WHEN posm."name" = 'PDB_2_Poster Display Board_BRAND_NOV_24' THEN mvp.amount ELSE 0 END,0)) "PDB_2_Poster Display Board_BRAND_NOV_24",
    --    SUM(COALESCE (CASE WHEN posm."name" = 'DLR_GPfi_DVC_Dec_24' THEN mvp.amount ELSE 0 END,0)) "DLR_GPfi_DVC_Dec_24" ,
        -- -- New Add -- -- 
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (219)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (219)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (618-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (618-629)_DATA_APR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (818-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (818-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (518-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (518-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (818-728)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (818-728)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (1099-629)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (1099-629)_DATA_APR_25",
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (518-529)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (518-529)_DATA_APR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Data (198)_DATA_APR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Data (198)_DATA_APR_25",
		-- --
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_MB_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "SS_MB_SKITTO_MAR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero Deno_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero Deno_SKITTO_MAR_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_M_RC_SKITTO_MAR_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_M_RC_SKITTO_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Telco_GA_MAR_25' THEN mvp.amount ELSE 0 END,0)) "FST_Telco_GA_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'COVS_Net_1_No._Network_DATA_MAR_25' THEN mvp.amount ELSE 0 END,0)) "COVS_Net_1_No._Network_DATA_MAR_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_Mar_25' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_Mar_25",
		-- --
	--	SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Flexi_BRND_Feb_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Flexi_BRND_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_GPfi_DVC_Feb_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_GPfi_DVC_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_4G Upgrade_NET_Feb_25' THEN mvp.amount ELSE 0 END,0)) "SS_4G Upgrade_NET_Feb_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_MyGP_App_ MYGP_MAR-25' THEN mvp.amount ELSE 0 END,0)) "SS_MyGP_App_ MYGP_MAR-25",
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "SS_1_No_Network_NET_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "FST_1_No_Network_NET_MAY_25"	,
		SUM(COALESCE(CASE WHEN posm."name" = 'SS_Net_1_No_Network_NET_MAY_25' THEN mvp.amount ELSE 0 END,0)) "SS_Net_1_No_Network_NET_MAY_25",	
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_39_Data_&_29_Voice_SC_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_39_Data_&_29_Voice_SC_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Flexiplan_MYGP_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_Flexiplan_MYGP_MAY_25",
		-- --
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Rate_Cutter_RC_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Rate_Cutter_RC_JUN_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_S_GA_198_Offer_SKITTO_MAY_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_S_GA_198_Offer_SKITTO_MAY_25",
		SUM(COALESCE(CASE WHEN posm."name" = 'FST_Haat_Bazar_NET_JUN_25' THEN mvp.amount ELSE 0 END,0)) "FST_Haat_Bazar_NET_JUN_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25",
		-- --
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Offer_BNDL_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Offer_BNDL_AUG_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_MyGP App_MYGP_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_MyGP App_MYGP_AUG_25",
		--SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Limitless_DATA_AUG_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Limitless_DATA_AUG_25",
        -- --
		SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_699_Bundle_Offer_BNDL_AUG 25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_699_Bundle_Offer_BNDL_AUG 25",
        -- --
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_SIM_&_Reload_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "FST_SIM_&_Reload_SKITTO_OCT_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Xtra_SKITTO_OCT_25' THEN mvp.amount ELSE 0 END,0)) "SS_Xtra_SKITTO_OCT_25",
        -- --
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bioscope+_CNT_Nov_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bioscope+_CNT_Nov_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'COVS_No._1_in_Internet_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "COVS_No._1_in_Internet_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Data_Hero_599_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Data_Hero_599_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Data_Hero_698_DATA_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Data_Hero_698_DATA_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Hero_699_BNDL_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Hero_699_BNDL_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Bundle_Hero_999_BNDL_NOV_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Bundle_Hero_999_BNDL_NOV_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_No.1_in_Internet_DATA_DEC_25' THEN mvp.amount ELSE 0 END,0)) "SS_No.1_in_Internet_DATA_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_49_Min_&_Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_49_Min_&_Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'ST_49_Min_&_Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "ST_49_Min_&_Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_39 Min_&_49 Data_SC_DEC_25' THEN mvp.amount ELSE 0 END,0)) "PSTR_39 Min_&_49 Data_SC_DEC_25",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_DEC_25' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_DEC_25",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Data_698_DATA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Data_698_DATA_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Data_499_DATA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Data_499_DATA_JAN_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Bundle_999_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Bundle_999_BNDL_JAN_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Hero_Bundle_498_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Hero_Bundle_498_BNDL_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_498_BNDL_JAN_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_498_BNDL_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_JAN_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_JAN_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_BNDL_FEB_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_BNDL_FEB_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Monthly_Bundle_BNDL_FEB_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Monthly_Bundle_BNDL_FEB_26",
        --SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_Ramadan_Poster__18th Feb_26_10X20' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_Ramadan_Poster__18th Feb_26_10X20",
        --SUM(COALESCE(CASE WHEN posm."name" = 'Ebadah_Sticker_9x6' THEN mvp.amount ELSE 0 END,0)) "Ebadah_Sticker_9x6",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Internet_e_1_no_Mar_26' THEN mvp.amount ELSE 0 END,0)) "SS_Internet_e_1_no_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Bondho_Simer_1_no_offer_Mar_26' THEN mvp.amount ELSE 0 END,0)) "FST_Bondho_Simer_1_no_offer_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'BNT_9”x11.5_Mar_26' THEN mvp.amount ELSE 0 END,0)) "BNT_9”x11.5_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Flexi_RLD_Mar_26' THEN mvp.amount ELSE 0 END,0)) "FST_Flexi_RLD_Mar_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Hajj_Roaming_Apr_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Hajj_Roaming_Apr_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_29 Min_&_49_Data_SC_Apr_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_29 Min_&_49_Data_SC_Apr_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'COVS_Net_1_No._Network_DATA_APR_26' THEN mvp.amount ELSE 0 END,0)) "COVS_Net_1_No._Network_DATA_APR_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Voice_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Voice_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Monthly_Bundle_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Monthly_Bundle_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_GPFI_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_GPFI_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_GPFI_Order_May_26' THEN mvp.amount ELSE 0 END,0)) "FST_GPFI_Order_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Monthly_Bundle_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_Monthly_Bundle_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_GPFI_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_GPFI_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Data_2_Deno_219_tk_&_599_tk_May_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Data_2_Deno_219_tk_&_599_tk_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Internet_e_1_no_May_26' THEN mvp.amount ELSE 0 END,0)) "SS_Internet_e_1_no_May_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Skitto_Recharge_June_26' THEN mvp.amount ELSE 0 END,0)) "FST_Skitto_Recharge_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_RC_1p_sec_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_RC_1p_sec_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_L700_MHz_metro_&_urban_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_L700_MHz_metro_&_urban_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_L700_MHz_semi_urban_&_rural_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_L700_MHz_semi_urban_&_rural_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_L700MHz_June_26' THEN mvp.amount ELSE 0 END,0)) "SS_L700MHz_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26' THEN mvp.amount ELSE 0 END,0)) "SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Voice_&_Data_Combo_Scratch_Card_June_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Voice_&_Data_Combo_Scratch_Card_June_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Power_Load_July_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Power_Load_July_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_Churn_Back_July_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_Churn_Back_July_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Skitto_Sim_Reload_August_26' THEN mvp.amount ELSE 0 END,0)) "FST_Skitto_Sim_Reload_August_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'PSTR_R_New_GA_Offer_SKITTO_August_26' THEN mvp.amount ELSE 0 END,0)) "PSTR_R_New_GA_Offer_SKITTO_August_26"
        SUM(COALESCE(CASE WHEN posm."name" = 'SS_Roaming_On_Aug_26' THEN mvp.amount ELSE 0 END,0)) "SS_Roaming_On_Aug_26",
        SUM(COALESCE(CASE WHEN posm."name" = 'FST_Roaming_On_Aug_26' THEN mvp.amount ELSE 0 END,0)) "FST_Roaming_On_Aug_26"
FROM
    gp.market_visit_posm_counts mvp
JOIN
    gp.materials posm ON mvp.posm_id = posm.id
WHERE
    mvp.visit_date BETWEEN '{to_date}' AND '{custom_date}'
GROUP BY
    mvp.visit_id
ORDER BY
    mvp.visit_id
)
SELECT
    mv.visit_id,
mv.visit_date,
ud. "Circle",
ud. "Region",
ud. "Cluster",
ud. "Territory",
ud. "Distribution House",
ud. "DH Code",
ud. "DH Lat-Long",
mv.route AS "Route",
mv.outlet_code AS "PoS Code",
mv.geo_checked_by AS "Geo Checked By",
mv.is_skitto AS "Is Skitto PoS",
ud. "ME Code", ud. "ME Name",
ud. "ME Contact No.",
ud. "ME Email",
mv.visit_lat_long AS "Visit Geo Location",
mv.operational_status AS "Visit Status",
sd.which_is_right_for_pos as "which_is_right_for_pos",
sd.pos_status as "Pos_Status",
mv.scheduled_visit AS "Scheduled Visit", 
case    
when sd.pre_execution_photo is null or  sd.pre_execution_photo = '' then ''
else concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.pre_execution_photo) 
end AS "Pre Execution",
sd.posm_exists AS "POSM Existance",
sd. "temporarily_&_permanently_closed_photo" AS " Temporary Or Permanently Closed",
sd. "not_found_&_moved_photo" AS "Not Found Or Moved",
sd.old_posm_list as"Old POSM List",
sd.do_posm_remove as "Do POSM Remove",
sd.old_posm_remove_list as"Old POSM Remove List", 
sd.old_posm_counts as "Old POSM Counts",
sd.removed_posm_counts as "Removed POSM Counts",
sd.sim_sell as "Sim_sell",
sd.gp_sim_sell as "Gp_sim_sell",
sd.other_sim_sell as "Other_sim_sell",
sd.finger_print_scanner as "Finger_print_scanner",
sd.other_company_scanner as "Other_company_scanner",
--pd."PSTR_L_GPfi_DVC_Dec_24",
--  pd."PSTR_R_1 No. Offer_DATA_NOV_24",
--  pd."PSTR_R_1 No. Plan_GA_NOV_24",
--pd."PSTR_R_Bondho Sim_CB_NOV_24",
--  pd."PSTR_R_Clear Cut_RC_NOV_24",
--  pd."PSTR_R_Limitless_DATA_NOV_24",
--  pd."PSTR_R_MyGP App_MYGP_NOV_24",
--pd."PSTR_S_106 Play Pack_PLP_NOV_24",
--pd."PSTR_S_49 & 29 Card_DATA_NOV_24",
--pd."PSTR_S_Toofan_PLP_NOV_24",
pd."FST_Sim Bikroy_GA_NOV_24",
--pd."FST_Internet POS_MYGP_NOV_24",
--pd."FST_Flexi_RLD_NOV_24",
pd."SS_Limitless_DATA_NOV_24",
pd."SS_ 1 No. Internet Network_NET_NOV_24",
--pd."SS_4G Upgrade_NET_NOV_24",
pd."SS_Apnar Elakay_NET_NOV_24",
pd."COVS_1 No. Network_NET_NOV_24",
pd."COVS_1 No. Network Internet_ DATA_NOV_24",
pd."PDB_4_Poster Display Board_BRAND_NOV_24",
pd."PDB_3_Poster Display Board_BRAND_NOV_24",
pd."PDB_2_Poster Display Board_BRAND_NOV_24",
--pd."DLR_GPfi_DVC_Dec_24",
--  pd."PSTR_R_1 No. Plan_GA_MAR_25",
--  pd."PSTR_R_Limitless_DATA_MAR_25",
--pd."PSTR_R_Bondho Sim_CB_MAR_25",
--pd."PSTR_R_Combo Offer_BNDL_MAR_25",
--pd."PSTR_R_MyGP App_MYGP_MAR_25",
pd."PSTR_R_Hero Data (219)_DATA_APR_25",
--pd."PSTR_R_Hero Data (618-629)_DATA_APR_25",
--  pd."PSTR_R_Hero Data (818-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (518-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (818-728)_DATA_APR_25",
--pd."PSTR_R_Hero Data (1099-629)_DATA_APR_25",
--pd."PSTR_R_Hero Data (518-529)_DATA_APR_25",
--pd."PSTR_R_Hero Data (198)_DATA_APR_25",
pd."SS_MB_SKITTO_MAR_25",
--  pd."PSTR_R_Hero Deno_SKITTO_MAR_25",
--  pd."PSTR_M_RC_SKITTO_MAR_25",
pd."FST_Telco_GA_MAR_25",
pd."COVS_Net_1_No._Network_DATA_MAR_25",
pd."FST_Flexi_RLD_Mar_25",
--pd."PSTR_R_Flexi_BRND_Feb_25",
pd."PSTR_R_GPfi_DVC_Feb_25",
pd."SS_4G Upgrade_NET_Feb_25",
pd."SS_MyGP_App_ MYGP_MAR-25",
pd."SS_1_No_Network_NET_MAY_25",
pd."FST_1_No_Network_NET_MAY_25",
pd."SS_Net_1_No_Network_NET_MAY_25",
--  pd."PSTR_39_Data_&_29_Voice_SC_MAY_25",
pd."PSTR_Flexiplan_MYGP_MAY_25",
pd."PSTR_R_Rate_Cutter_RC_JUN_25",
pd."PSTR_S_GA_198_Offer_SKITTO_MAY_25",
pd."FST_Haat_Bazar_NET_JUN_25",
--  pd."PSTR_R_OJ_Set_1_DATA_BNDL_JUN_25",
--  pd."PSTR_R_OJ_Set_2_DATA_BNDL_JUN_25",
-- pd."PSTR_R_Bundle_Offer_BNDL_AUG_25",
-- pd."PSTR_R_MyGP App_MYGP_AUG_25",
-- pd."PSTR_R_Limitless_DATA_AUG_25",
pd."PSTR_R_699_Bundle_Offer_BNDL_AUG 25",
-- pd."PSTR_R_Hero_Deno_Xtra_SKITTO_OCT_25",
pd."FST_SIM_&_Reload_SKITTO_OCT_25",
pd."SS_Xtra_SKITTO_OCT_25",
pd."PSTR_R_Bioscope+_CNT_Nov_25",
pd."COVS_No._1_in_Internet_DATA_NOV_25",
pd."PSTR_R_Data_Hero_599_DATA_NOV_25",
pd."PSTR_R_Data_Hero_698_DATA_NOV_25",
pd."PSTR_R_Bundle_Hero_699_BNDL_NOV_25",
pd."PSTR_R_Bundle_Hero_999_BNDL_NOV_25",
pd."SS_No.1_in_Internet_DATA_DEC_25",
pd."PSTR_49_Min_&_Data_SC_DEC_25",
pd."ST_49_Min_&_Data_SC_DEC_25",
pd."PSTR_39 Min_&_49 Data_SC_DEC_25",
pd."FST_Flexi_RLD_DEC_25",
--pd."PSTR_R_Hero_Data_698_DATA_JAN_26",
pd."PSTR_R_Hero_Data_499_DATA_JAN_26",
--pd."PSTR_R_Hero_Bundle_999_BNDL_JAN_26",
--pd."PSTR_R_Hero_Bundle_498_BNDL_JAN_26",
pd."PSTR_R_New_Sim_4GB+60_Mins_GA_JAN_26",
pd."SS_Monthly_Bundle_498_BNDL_JAN_26",
pd."PSTR_R_New_GA_Offer_SKITTO_JAN_26",
pd."SS_Monthly_Bundle_BNDL_FEB_26",
--pd."PSTR_Monthly_Bundle_BNDL_FEB_26",
-- pd."PSTR_R_Ramadan_Poster__18th Feb_26_10X20",
-- pd."Ebadah_Sticker_9x6",
pd."SS_Internet_e_1_no_Mar_26",
pd."FST_Bondho_Simer_1_no_offer_Mar_26",
pd."BNT_9”x11.5_Mar_26",
pd."FST_Flexi_RLD_Mar_26",
pd."PSTR_Hajj_Roaming_Apr_26",
pd."PSTR_29 Min_&_49_Data_SC_Apr_26",
pd."COVS_Net_1_No._Network_DATA_APR_26",
pd."PSTR_Voice_May_26",
pd."PSTR_Monthly_Bundle_May_26",
pd."SS_GPFI_May_26",
pd."FST_GPFI_Order_May_26",
pd."SS_Monthly_Bundle_May_26",
pd."PSTR_GPFI_May_26",
pd."PSTR_Data_2_Deno_219_tk_&_599_tk_May_26",
pd."SS_Internet_e_1_no_May_26",
pd."PSTR_R_New_GA_Offer_SKITTO_June_26",
pd."FST_Skitto_Recharge_June_26",
pd."PSTR_RC_1p_sec_June_26",
pd."PSTR_BIOSCOPE_FIFA_World_Cup_Fixture_June_26",
pd."PSTR_L700_MHz_metro_&_urban_June_26",
pd."PSTR_L700_MHz_semi_urban_&_rural_June_26",
pd."SS_L700MHz_June_26",
pd."SS_RC_1p_sec_64tk_&_94tk_Voice_Recharge_June_26",
pd."PSTR_Voice_&_Data_Combo_Scratch_Card_June_26",
pd."PSTR_Power_Load_July_26",
pd."PSTR_Churn_Back_July_26",
pd."FST_Skitto_Sim_Reload_August_26",
pd."PSTR_R_New_GA_Offer_SKITTO_August_26",
pd."SS_Roaming_On_Aug_26",
pd."FST_Roaming_On_Aug_26",
case    
when sd.execution_photo_center is null or  sd.execution_photo_center = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_center) 
end AS "Execution Center",
case    
when sd.execution_photo_left is null or  sd.execution_photo_left = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_left) 
end AS "Execution Left",
case    
when sd.execution_photo_right is null or  sd.execution_photo_right = '' then ''
else
concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/',sd.execution_photo_right) 
end AS "Execution Right",
sd.new_posm_counts as "New POSM Counts",
case    
when sd.post_execution_photo is null or  sd.post_execution_photo = '' then ''
else concat('https://gp-web-uploads.s3.ap-southeast-1.amazonaws.com/Production/gp/Images/Retail/MarketVisit/',TO_CHAR(mv.visit_date, 'YYYYMM'),'/', sd.post_execution_photo) 
end AS "Post Execution",
sd.new_posm_selection as "New POSM Selection",
sd.pos_business_type as "Pos Business Type",
sd.pos_structure as "POS Structure",
sd.pos_location as "POS Place",
sd.gp_fascia as "GP Fascia",
sd.gp_fascia_type AS "GP Fascia Type",
sd.other_fascia AS "Other Fascia",
sd.which_other_fascia AS "Other Fascia List",
sd.business_other_operator as "business_other_operator",
--sd.theamatic as "theamatic",
--sd.themetic_puzzle_block as "themetic_puzzle_block",
sd.poster_counts as "poster_counts"
--sd."10x20_poster_counts" as "10x20_poster_counts"
FROM
    mv_data mv
    LEFT JOIN user_data ud ON mv.me = ud."ME Code"
    LEFT JOIN survey_data sd ON mv.visit_id = sd.visit_id
    LEFT JOIN posm_data pd ON mv.visit_id = pd.visit_id
;
""")

def fetch_raw_data(to_date, custom_date, query_function):
    try:
        query = query_function(to_date, custom_date)
        df = pd.read_sql_query(query, con=engine)  
        return df
    except Exception as e:
        print(f"❌ Error fetching data: {e}")
        return pd.DataFrame()

def get_day_with_suffix(day):
    if 11 <= day <= 13:
        return f"{day}th"
    last_digit = day % 10
    if last_digit == 1:
        return f"{day}st"
    elif last_digit == 2:
        return f"{day}nd"
    elif last_digit == 3:
        return f"{day}rd"
    else:
        return f"{day}th"

def format_date_for_file_name(start_date_str, end_date_str):
    start = datetime.strptime(start_date_str, "%Y-%m-%d")
    end = datetime.strptime(end_date_str, "%Y-%m-%d")
    start_day = get_day_with_suffix(start.day)
    start_month = start.strftime("%B")
    end_day = get_day_with_suffix(end.day)
    end_month = end.strftime("%B")
    if start.date() == end.date():
        return f"{start_day} {start_month}"
    return f"{start_day} {start_month} to {end_day} {end_month}"

def format_sheet_name(to_date, custom_date):
    start = datetime.strptime(to_date, '%Y-%m-%d')
    end = datetime.strptime(custom_date, '%Y-%m-%d')

    if start.date() == end.date():
        return f"{get_day_with_suffix(end.day)} {end.strftime('%B')} -{end.year}"
    if start.year == end.year and start.month == end.month:
        return f"{get_day_with_suffix(start.day)} to {get_day_with_suffix(end.day)} {end.strftime('%B')} -{end.year}"
    return f"{get_day_with_suffix(start.day)} {start.strftime('%B')} to {get_day_with_suffix(end.day)} {end.strftime('%B')} -{end.year}"

def is_file_locked(file_path):
    """
    Best-effort check for whether file_path is currently open/locked by
    another program (most commonly: the report .xlsx is open in Excel).
    Opening for exclusive read+write fails with PermissionError/OSError on
    Windows if another process holds it open — this catches that BEFORE
    openpyxl tries to append and blows up with a cryptic 'Bad CRC-32' error.
    """
    if not os.path.exists(file_path):
        return False
    try:
        with open(file_path, 'r+b'):
            pass
        return False
    except (PermissionError, OSError):
        return True


def sheet_exists(file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        return sheet_name in wb.sheetnames
    except Exception as e:
        print(f"❌ Error checking sheet existence: {e}")
        return False

def format_excel_file(excel_file_path, sheet_name=None):
    try:
        wb = load_workbook(excel_file_path)
        if sheet_name and sheet_name in wb.sheetnames:
            sheets_to_format = [wb[sheet_name]]
        else:
            sheets_to_format = wb.worksheets

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for ws in sheets_to_format:
            if ws.title == "Sheet" and len(wb.sheetnames) > 1:
                continue

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None:
                            val_len = len(str(cell.value))
                            if val_len > max_length:
                                max_length = val_len
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])
        wb.save(excel_file_path)
        print(f"✅ Excel formatted successfully: {os.path.basename(excel_file_path)}")
    except Exception as e:
        print(f"❌ Error formatting Excel file: {e}")


def get_or_rename_folder(base_path, start_date, end_date):
    new_range = format_date_for_file_name(start_date, end_date)
    new_folder = f"POS Daily Report ({new_range})"
    new_path = os.path.join(base_path, new_folder)

    for folder in os.listdir(base_path):
        if folder.startswith("POS Daily Report") and os.path.isdir(os.path.join(base_path, folder)):
            old_path = os.path.join(base_path, folder)
            if new_folder != folder:
                try:
                    os.rename(old_path, new_path)
                    print(f"♻️ Renamed folder: {folder} ➞ {new_folder}")
                    return new_path
                except Exception as e:
                    print(f"❌ Error renaming folder: {e}")
                    return old_path
            else:
                return old_path

    os.makedirs(new_path, exist_ok=True)
    print(f"✅ Created new folder: {new_folder}")
    return new_path


def cleanup_old_circle_files(folder_path, circle, keep_file_name):
    """Clean up old file(s) for a given circle in folder_path if the filename has changed."""
    prefix = f"{circle}_POS Daily Report"
    for fname in os.listdir(folder_path):
        if fname.startswith(prefix) and fname.endswith(".xlsx") and fname != keep_file_name:
            old_file_path = os.path.join(folder_path, fname)
            try:
                os.remove(old_file_path)
                print(f"🗑️ Removed outdated file for {circle}: {fname}")
            except Exception as e:
                print(f"⚠️ Could not remove old file {fname}: {e}")


def save_the_data(to_date_param=None, custom_date_param=None, log=print, base_folder=None):
    circle_files = {}

    start_dt = to_date_param if to_date_param else globals().get('to_date', '2026-07-25')
    end_dt = custom_date_param if custom_date_param else globals().get('custom_date', (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d'))

    df = fetch_raw_data(start_dt, end_dt, pos_daily_rpt)
    if df.empty:
        log("❌ No data fetched from the database.")
        return circle_files

    if not base_folder:
        base_folder = r'E:/GP/Report/POS Daily Report/Daily POS Visit Reports'
    folder_path = get_or_rename_folder(base_folder, start_dt, end_dt)

    # Convert visit_date to string date YYYY-MM-DD for grouping
    try:
        df['visit_date_str'] = pd.to_datetime(df['visit_date']).dt.strftime('%Y-%m-%d')
    except Exception as e:
        log(f"⚠️ Error formatting visit_date: {e}")
        df['visit_date_str'] = df['visit_date'].astype(str)

    new_range = format_date_for_file_name(start_dt, end_dt)

    for circle in df['Circle'].unique():
        circle_df = df[df['Circle'] == circle]

        new_file_name = f"{circle}_POS Daily Report ({new_range}).xlsx"
        file_path = os.path.join(folder_path, new_file_name)

        cleanup_old_circle_files(folder_path, circle, new_file_name)

        if os.path.exists(file_path) and is_file_locked(file_path):
            log(f"❌ Skipping {circle}: '{os.path.basename(file_path)}' is currently open in Excel "
                f"(or locked by another program). Close the file and click 'Download Report' again.")
            continue

        circle_files[circle] = file_path

        # Group and write date-wise in chronological order into a single Excel file
        unique_dates = sorted(circle_df['visit_date_str'].unique())

        try:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                for dt_str in unique_dates:
                    day_df = circle_df[circle_df['visit_date_str'] == dt_str].copy()
                    if 'visit_date_str' in day_df.columns:
                        day_df = day_df.drop(columns=['visit_date_str'])

                    sheet_name = format_sheet_name(dt_str, dt_str)
                    day_df.to_excel(writer, index=False, sheet_name=sheet_name)

            format_excel_file(file_path)
            log(f"✅ Created report for {circle} with {len(unique_dates)} daily sheet(s) ({', '.join(unique_dates)})")
        except zipfile.BadZipFile:
            log(f"❌ {circle}: '{os.path.basename(file_path)}' is corrupted or was open in Excel during the "
                f"write (Bad CRC-32). Close the file in Excel, delete it if needed, and re-run.")
        except PermissionError:
            log(f"❌ {circle}: permission denied writing '{os.path.basename(file_path)}' — "
                f"it's likely open in Excel or another program. Close it and re-run.")
        except Exception as e:
            log(f"❌ Error writing report for {circle}: {e}")

    return circle_files


def zip_files(folder_path, zip_output_path):
    try:
        with zipfile.ZipFile(zip_output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, folder_path)
                    zipf.write(file_path, arcname)
        print(f"✅ Folder zipped at {zip_output_path}")
    except Exception as e:
        print(f"❌ Error zipping files: {e}")

def load_email_recipients(xlsx_path, log=print):
    """
    Reads the uploaded recipient-mapping Excel file.
    One worksheet per Circle. Columns: 'To', 'Cc'. Every non-empty cell in each
    column across all rows is collected (multiple recipients per circle).

    Returns: { "CIRCLE_NAME_UPPER": {"to": [...], "cc": [...]}, ... }
    """
    mapping = {}
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        log(f"Could not read recipient list file: {e}")
        return mapping

    for ws in wb.worksheets:
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        try:
            to_col = headers.index("to")
        except ValueError:
            to_col = None
        try:
            cc_col = headers.index("cc")
        except ValueError:
            cc_col = None

        to_list, cc_list = [], []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if to_col is not None and to_col < len(row) and row[to_col]:
                val = str(row[to_col]).strip()
                if val and val not in to_list:
                    to_list.append(val)
            if cc_col is not None and cc_col < len(row) and row[cc_col]:
                val = str(row[cc_col]).strip()
                if val and val not in cc_list:
                    cc_list.append(val)

        circle_key = ws.title.strip().upper()
        mapping[circle_key] = {"to": to_list, "cc": cc_list}
        log(f"Loaded {len(to_list)} To / {len(cc_list)} Cc recipients for circle '{ws.title}'")

    return mapping


def send_circle_email(circle, file_path, to_emails, cc_emails, formatted_range, log=print):
    """
    Sends ONE email for ONE circle's report file.
    - Uploads that single file to Google Drive and includes the share link.
    - ALSO attaches the .xlsx file directly to the email.
    """
    from_email = "shadman.sayeid@v2.ltd"
    email_password = "ujmt gpgi ctbx dohe"
    group_mail = "cockpit.glm@v2.ltd"
    subject = f"{circle} POS Daily Report ({formatted_range})"

    if not to_emails:
        log(f"Skipping {circle}: no To recipients found in the uploaded mapping file.")
        return False

    if not os.path.exists(file_path):
        log(f"Skipping {circle}: file not found at {file_path}")
        return False

    try:
        from pydrive2.auth import GoogleAuth
        from pydrive2.drive import GoogleDrive

        gauth = GoogleAuth()
        gauth.LoadClientConfigFile("client_secrets.json")
        gauth.settings['get_refresh_token'] = True
        gauth.settings['oauth_scope'] = ['https://www.googleapis.com/auth/drive']
        gauth.LoadCredentialsFile("mycreds.txt")
        if gauth.credentials is None:
            gauth.LocalWebserverAuth()
        elif gauth.access_token_expired:
            gauth.Refresh()
        else:
            gauth.Authorize()
        gauth.SaveCredentialsFile("mycreds.txt")

        drive = GoogleDrive(gauth)

        drive_link = None
        try:
            file_drive = drive.CreateFile({'title': os.path.basename(file_path)})
            file_drive.SetContentFile(file_path)
            file_drive.Upload()
            file_drive.InsertPermission({'type': 'anyone', 'value': 'anyone', 'role': 'reader'})
            drive_link = file_drive['alternateLink']
        except Exception as e:
            log(f"Drive upload failed for {circle} (continuing with attachment only): {e}")

        disclaimer_html = (
            "<hr>"
            "<small> Disclaimer: The information in this document and attachments is confidential and may also be legally privileged. It is intended only for the use of the named recipient. Internet communications are not secure and therefore, Madly Ltd. does not accept legal responsibility for the contents of this message. If you are not the intended recipient, please notify us immediately and then delete this document. Do not disclose the contents of this document to any other person, nor take any copies. Violation of this notice may be unlawful."
            "</small>"
        )
        link_html = f"<p><a href=\"{drive_link}\">{drive_link}</a></p>" if drive_link else ""
        body = (
            f"<p>Dear Concern,</p>"
            f"<p>Please find attached the <b>{circle} POS Daily Report ({formatted_range})</b>.</p>"
            f"{link_html}"
            f"<p><i>[Please note, This is an automated email report. Please do not reply to this email or mark it as spam.]</i><br>"
            f"Regards,<br><b>Cockpit GLM System</b></p>"
            f"{disclaimer_html}"
        )

        msg = MIMEMultipart()
        msg['From'] = group_mail
        msg['Subject'] = subject
        msg['To'] = ", ".join(to_emails)
        if cc_emails:
            msg['Cc'] = ", ".join(cc_emails)
        all_recipients = to_emails + cc_emails
        msg.attach(MIMEText(body, 'html'))

        max_gmail_attachment_size = 24 * 1024 * 1024  # 24 MB safe threshold
        attachment_sent = False
        if os.path.getsize(file_path) <= max_gmail_attachment_size:
            with open(file_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(file_path)}"')
            msg.attach(part)
            attachment_sent = True
        else:
            log(f"[{circle}] File is too large for Gmail attachment ({os.path.getsize(file_path) / (1024*1024):.1f} MB). Sending Drive link only.")
            body = (
                f"<p>Dear Concern,</p>"
                 f"<p>Please find attached the <b>{circle} POS Daily Report ({formatted_range})</b>.</p>"
                f"<p>The file is too large to attach via Gmail, so it is available only through the Google Drive link below.</p>"
                f"{link_html}"
                f"<p><i>[Please note, This is an automated email report. Please do not reply to this email or mark it as spam.]</i><br>"
                f"Regards,<br><b>Cockpit GLM System</b></p>"
                f"{disclaimer_html}"
            )
            msg = MIMEMultipart()
            msg['From'] = group_mail
            msg['Subject'] = subject
            msg['To'] = ", ".join(to_emails)
            if cc_emails:
                msg['Cc'] = ", ".join(cc_emails)
            msg.attach(MIMEText(body, 'html'))

        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(from_email, email_password)
            server.sendmail(group_mail, all_recipients, msg.as_string())

        if attachment_sent:
            log(f"[{circle}] Email sent to {', '.join(all_recipients)}")
        else:
            log(f"[{circle}] Email sent to {', '.join(all_recipients)} with Drive link only (attachment skipped due to size) ")
        return True
    except Exception as e:
        log(f"[{circle}] Error sending email: {e}")
        return False


def generate_and_load(email_list_path, custom_date_val=None, to_date_val=None, log=print, base_folder=None):
    """
    Step 1 of the pipeline: fetch data, write per-circle files (into base_folder
    if given, else the default hardcoded path), load recipients.
    Does NOT send anything. Returns everything needed for a preview screen.
    """
    global custom_date, to_date

    if custom_date_val:
        custom_date = custom_date_val
    else:
        custom_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')

    if to_date_val:
        to_date = to_date_val
    else:
        to_date = '2026-07-25'

    formatted_range = format_date_for_file_name(to_date, custom_date)

    log(f"🚀 Starting run for {custom_date} (range: {formatted_range})")

    circle_files = save_the_data(to_date_param=to_date, custom_date_param=custom_date, log=log, base_folder=base_folder)
    if not circle_files:
        log("❌ No circle files were generated. Stopping.")
        return {}, {}, formatted_range

    recipients_map = load_email_recipients(email_list_path, log=log)
    if not recipients_map:
        log("❌ No recipients loaded from the mapping file. Stopping.")
        return circle_files, {}, formatted_range

    return circle_files, recipients_map, formatted_range


def send_selected(circle_files, recipients_map, formatted_range, selected_circles, log=print):
    """
    Step 2 of the pipeline: send emails only for circles in `selected_circles`
    (a set/list of circle names as they appear in circle_files' keys).
    """
    sent, skipped = 0, 0
    for circle, file_path in circle_files.items():
        if circle not in selected_circles:
            log(f"⏭️ Skipped by user: {circle}")
            skipped += 1
            continue

        entry = recipients_map.get(str(circle).strip().upper())
        if not entry:
            log(f"⚠️ No recipient entry found for circle '{circle}' in the uploaded file. Skipping.")
            skipped += 1
            continue

        ok = send_circle_email(
            circle=circle,
            file_path=file_path,
            to_emails=entry.get("to", []),
            cc_emails=entry.get("cc", []),
            formatted_range=formatted_range,
            log=log,
        )
        sent += 1 if ok else 0
        skipped += 0 if ok else 1

    log(f"🏁 Done. Emails sent: {sent}, skipped: {skipped}")


def run_pipeline(email_list_path, log=print):
    """Convenience wrapper: generate, then send everything with a match — no preview.
    Kept for scripts/automation that don't need the GUI preview step."""
    circle_files, recipients_map, formatted_range = generate_and_load(email_list_path, log=log)
    if not circle_files or not recipients_map:
        return
    send_selected(circle_files, recipients_map, formatted_range, set(circle_files.keys()), log=log)


class ReportSenderGUI:
    def __init__(self, root):
        self.root = root
        root.title("POS Daily Report — Download & Send")
        root.geometry("780x620")
        root.configure(bg="#f8f9fa")

        self.circle_files = {}
        self.recipients_map = {}
        self.formatted_range = ""
        self.include_vars = {}
        self.auto_send_after_generate = False
        self.is_loading = False

        # Custom modern styling configurations
        self.bg_color = "#f8f9fa"
        self.primary_color = "#0d6efd"
        self.primary_hover = "#0b5ed7"
        self.success_color = "#198754"
        self.success_hover = "#157347"
        self.btn_bg_color = "#e2e6ea"
        self.btn_hover_color = "#dae0e5"
        self.text_color = "#212529"

        # --- Recipient mapping excel ---
        tk.Label(root, text="Recipient mapping Excel file (one sheet per Circle, columns To/Cc):", bg=self.bg_color, fg=self.text_color, font=("Segoe UI", 9, "bold")).pack(
            anchor="w", padx=10, pady=(10, 0)
        )
        path_frame = tk.Frame(root, bg=self.bg_color)
        path_frame.pack(fill="x", padx=10, pady=5)
        self.path_var = tk.StringVar()
        self.path_entry = tk.Entry(path_frame, textvariable=self.path_var)
        self.path_entry.pack(side="left", fill="x", expand=True, ipady=3)
        self.browse_file_btn = tk.Button(path_frame, text="Browse & Upload...", command=self.browse_email_file, bg=self.btn_bg_color, relief="flat", activebackground=self.btn_hover_color)
        self.browse_file_btn.pack(side="left", padx=(5, 0))
        self.bind_hover(self.browse_file_btn, self.btn_bg_color, self.btn_hover_color)

        # --- Destination folder for downloaded reports ---
        tk.Label(root, text="Destination folder to save reports on this PC:", bg=self.bg_color, fg=self.text_color, font=("Segoe UI", 9, "bold")).pack(
            anchor="w", padx=10, pady=(8, 0)
        )
        dest_frame = tk.Frame(root, bg=self.bg_color)
        dest_frame.pack(fill="x", padx=10, pady=5)
        self.dest_var = tk.StringVar()
        self.dest_entry = tk.Entry(dest_frame, textvariable=self.dest_var)
        self.dest_entry.pack(side="left", fill="x", expand=True, ipady=3)
        self.browse_dest_btn = tk.Button(dest_frame, text="Choose Folder...", command=self.browse_dest_folder, bg=self.btn_bg_color, relief="flat", activebackground=self.btn_hover_color)
        self.browse_dest_btn.pack(side="left", padx=(5, 0))
        self.bind_hover(self.browse_dest_btn, self.btn_bg_color, self.btn_hover_color)

        # --- Date settings frame ---
        tk.Label(root, text="Report Date Range (YYYY-MM-DD):", bg=self.bg_color, fg=self.text_color, font=("Segoe UI", 9, "bold")).pack(
            anchor="w", padx=10, pady=(8, 0)
        )
        date_frame = tk.Frame(root, bg=self.bg_color)
        date_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(date_frame, text="Start Date:", bg=self.bg_color, fg=self.text_color).pack(side="left")
        self.start_date_var = tk.StringVar(value="2026-07-25")
        self.start_entry = tk.Entry(date_frame, textvariable=self.start_date_var, width=15)
        self.start_entry.pack(side="left", padx=(5, 15), ipady=2)
        
        tk.Label(date_frame, text="End Date:", bg=self.bg_color, fg=self.text_color).pack(side="left")
        yesterday_str = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        self.end_date_var = tk.StringVar(value=yesterday_str)
        self.end_entry = tk.Entry(date_frame, textvariable=self.end_date_var, width=15)
        self.end_entry.pack(side="left", padx=(5, 0), ipady=2)

        self.status_label = tk.Label(root, text="Status: waiting for recipient file + destination folder.", anchor="w", fg="#6c757d", bg=self.bg_color, font=("Segoe UI", 9, "italic"))
        self.status_label.pack(fill="x", padx=10, pady=(4, 0))

        # --- Preview area (populated after generation) ---
        self.preview_label = tk.Label(root, text="Preview: (nothing generated yet)", anchor="w", bg=self.bg_color, fg=self.text_color)
        self.preview_label.pack(fill="x", padx=10, pady=(8, 0))

        self.preview_frame = tk.Frame(root, relief="groove", borderwidth=1, bg=self.bg_color)
        self.preview_frame.pack(fill="x", padx=10, pady=(5, 5))

        # --- Action buttons ---
        btn_frame = tk.Frame(root, bg=self.bg_color)
        btn_frame.pack(fill="x", padx=10, pady=(0, 10))

        self.generate_btn = tk.Button(
            btn_frame, text="Download Report", command=lambda: self.start_generate(auto_send=False),
            bg=self.primary_color, fg="white", height=2, relief="flat", activebackground=self.primary_hover, font=("Segoe UI", 10, "bold")
        )
        self.generate_btn.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.bind_hover(self.generate_btn, self.primary_color, self.primary_hover)

        self.confirm_btn = tk.Button(
            btn_frame, text="Download and Send", command=self.download_and_send,
            bg=self.success_color, fg="white", height=2, relief="flat", activebackground=self.success_hover, font=("Segoe UI", 10, "bold")
        )
        self.confirm_btn.pack(side="left", fill="x", expand=True, padx=(5, 0))
        self.bind_hover(self.confirm_btn, self.success_color, self.success_hover)

        tk.Label(root, text="Log:", bg=self.bg_color, fg=self.text_color, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10)
        self.log_box = scrolledtext.ScrolledText(root, height=12, state="disabled", font=("Consolas", 9))
        self.log_box.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def bind_hover(self, widget, normal_bg, hover_bg):
        widget.bind("<Enter>", lambda e: widget.config(bg=hover_bg) if str(widget['state']) != 'disabled' else None)
        widget.bind("<Leave>", lambda e: widget.config(bg=normal_bg) if str(widget['state']) != 'disabled' else None)

    def animate_status(self, base_text):
        chars = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
        def _step(idx=0):
            if self.is_loading:
                self.status_label.config(text=f"Status: {base_text} {chars[idx % len(chars)]}", fg="#0d6efd")
                self.root.after(100, _step, idx + 1)
        _step()

    # ---------- file / folder pickers ----------

    def browse_email_file(self):
        path = filedialog.askopenfilename(
            title="Select recipient mapping Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not path:
            return
        self.path_var.set(path)

        if not self.dest_var.get().strip():
            self.status_label.config(text="Status: recipient file loaded. Please choose a destination folder.")
        else:
            self.status_label.config(text="Status: ready. Click 'Download Report' or 'Download and Send' to start.")

    def browse_dest_folder(self):
        folder = filedialog.askdirectory(title="Choose destination folder for reports")
        if folder:
            self.dest_var.set(folder)
            if self.path_var.get().strip():
                self.status_label.config(text="Status: ready. Click 'Download Report' or 'Download and Send' to start.")
            else:
                self.status_label.config(text="Status: destination folder loaded. Please choose a recipient mapping Excel file.")

    # ---------- logging / helpers ----------

    def log(self, message):
        def _write():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", str(message) + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.root.after(0, _write)

    def open_file(self, file_path):
        try:
            os.startfile(file_path)  # Windows only
        except AttributeError:
            self.log(f"Can't auto-open on this OS. File is at: {file_path}")
        except Exception as e:
            self.log(f"Could not open file: {e}")

    def build_preview(self):
        for widget in self.preview_frame.winfo_children():
            widget.destroy()
        self.include_vars = {}

        if not self.circle_files:
            self.preview_label.config(text="Preview: no circle files were generated.")
            return

        header = tk.Frame(self.preview_frame, bg=self.bg_color)
        header.pack(fill="x", padx=5, pady=(5, 2))
        tk.Label(header, text="Send?", width=6, anchor="w", font=("Segoe UI", 9, "bold"), bg=self.bg_color).pack(side="left")
        tk.Label(header, text="Circle", width=16, anchor="w", font=("Segoe UI", 9, "bold"), bg=self.bg_color).pack(side="left")
        tk.Label(header, text="To / Cc", width=10, anchor="w", font=("Segoe UI", 9, "bold"), bg=self.bg_color).pack(side="left")
        tk.Label(header, text="File", anchor="w", font=("Segoe UI", 9, "bold"), bg=self.bg_color).pack(side="left", fill="x", expand=True)

        for circle, file_path in self.circle_files.items():
            entry = self.recipients_map.get(str(circle).strip().upper(), {})
            to_count = len(entry.get("to", []))
            cc_count = len(entry.get("cc", []))
            has_recipients = to_count > 0

            row = tk.Frame(self.preview_frame, bg=self.bg_color)
            row.pack(fill="x", padx=5, pady=2)

            var = tk.BooleanVar(value=has_recipients)
            self.include_vars[circle] = var
            cb = tk.Checkbutton(row, variable=var, width=4, bg=self.bg_color, activebackground=self.bg_color)
            cb.pack(side="left")
            if not has_recipients:
                cb.config(state="disabled")

            label_color = "black" if has_recipients else "red"
            tk.Label(row, text=circle, width=16, anchor="w", fg=label_color, bg=self.bg_color).pack(side="left")
            tk.Label(row, text=f"{to_count} / {cc_count}", width=10, anchor="w", fg=label_color, bg=self.bg_color).pack(side="left")
            tk.Label(row, text=os.path.basename(file_path), anchor="w", bg=self.bg_color).pack(side="left", fill="x", expand=True)
            
            open_btn = tk.Button(row, text="Open file", command=lambda p=file_path: self.open_file(p), bg=self.btn_bg_color, relief="flat", activebackground=self.btn_hover_color)
            open_btn.pack(side="right")
            self.bind_hover(open_btn, self.btn_bg_color, self.btn_hover_color)

        missing = [c for c, v in self.include_vars.items() if not v.get()]
        note = " (red rows have no recipients and are unchecked — fix the mapping file if needed)" if missing else ""
        self.preview_label.config(
            text=f"Preview: {len(self.circle_files)} circle file(s) generated.{note}"
        )

    # ---------- pipeline actions ----------

    def download_and_send(self):
        """'Download and Send' button: generates (if not already done) then sends after confirmation."""
        if not self.circle_files:
            self.auto_send_after_generate = True
            self.start_generate(auto_send=True)
        else:
            self.start_send()

    def start_generate(self, auto_send=False):
        email_list_path = self.path_var.get().strip()
        dest_folder = self.dest_var.get().strip()
        start_date_val = self.start_date_var.get().strip()
        end_date_val = self.end_date_var.get().strip()

        if not email_list_path or not os.path.exists(email_list_path):
            messagebox.showerror("Missing file", "Please select a valid recipient mapping Excel file first.")
            return
        if not dest_folder:
            messagebox.showerror("Missing folder", "Please choose a destination folder to save the reports first.")
            return
        if not os.path.exists(dest_folder):
            try:
                os.makedirs(dest_folder, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Invalid folder", f"Could not use that destination folder: {e}")
                return

        # Validate date formats (YYYY-MM-DD)
        date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        if not date_pattern.match(start_date_val) or not date_pattern.match(end_date_val):
            messagebox.showerror("Invalid Date Format", "Please enter dates in YYYY-MM-DD format.")
            return

        self.generate_btn.config(state="disabled", bg=self.btn_bg_color, fg="#6c757d")
        self.confirm_btn.config(state="disabled", bg=self.btn_bg_color, fg="#6c757d")
        self.is_loading = True
        self.animate_status("fetching data and generating reports")

        def worker():
            try:
                circle_files, recipients_map, formatted_range = generate_and_load(
                    email_list_path, 
                    custom_date_val=end_date_val, 
                    to_date_val=start_date_val, 
                    log=self.log, 
                    base_folder=dest_folder
                )
                self.circle_files = circle_files
                self.recipients_map = recipients_map
                self.formatted_range = formatted_range
                self.root.after(0, self.build_preview)
                self.root.after(0, lambda: self.status_label.config(
                    text=f"Status: {len(circle_files)} report(s) downloaded to {dest_folder}",
                    fg="#198754"
                ))
            except Exception as e:
                self.log(f"❌ Unexpected error during generation: {e}")
                self.root.after(0, lambda: self.status_label.config(text="Status: error occurred during generation.", fg="#dc3545"))
            finally:
                self.is_loading = False
                self.root.after(0, lambda: self.generate_btn.config(state="normal", bg=self.primary_color, fg="white"))
                self.root.after(0, lambda: self.confirm_btn.config(state="normal", bg=self.success_color, fg="white"))
                if auto_send and self.auto_send_after_generate:
                    self.auto_send_after_generate = False
                    self.root.after(0, self.start_send)

        threading.Thread(target=worker, daemon=True).start()

    def start_send(self):
        if not self.circle_files:
            messagebox.showwarning("Nothing to send", "Download the report(s) first.")
            return

        selected = [c for c, v in self.include_vars.items() if v.get()]
        if not selected:
            messagebox.showwarning("Nothing selected", "No circles are checked for sending.")
            return

        if not messagebox.askyesno(
            "Confirm send",
            f"Send emails for {len(selected)} circle(s): {', '.join(selected)}?"
        ):
            return

        self.confirm_btn.config(state="disabled", bg=self.btn_bg_color, fg="#6c757d")
        self.generate_btn.config(state="disabled", bg=self.btn_bg_color, fg="#6c757d")
        self.is_loading = True
        self.animate_status("sending emails")

        def worker():
            try:
                send_selected(
                    self.circle_files, self.recipients_map, self.formatted_range,
                    selected_circles=set(selected), log=self.log
                )
                self.root.after(0, lambda: self.status_label.config(text="Status: sending complete — check log for details.", fg="#198754"))
            except Exception as e:
                self.log(f"❌ Unexpected error while sending: {e}")
                self.root.after(0, lambda: self.status_label.config(text="Status: error occurred while sending.", fg="#dc3545"))
            finally:
                self.is_loading = False
                self.root.after(0, lambda: self.confirm_btn.config(state="normal", bg=self.success_color, fg="white"))
                self.root.after(0, lambda: self.generate_btn.config(state="normal", bg=self.primary_color, fg="white"))

        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    root = tk.Tk()
    app = ReportSenderGUI(root)
    root.mainloop()

    